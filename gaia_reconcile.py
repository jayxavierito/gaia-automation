from __future__ import annotations

import argparse
import csv
import json
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

from gaia_converter import clean_text, normalize_match_text, normalize_unit


@dataclass
class GaiaReviewRow:
    queue: str
    workbook: str
    gaia_row: int
    name: str
    specification: str
    unit: str
    price: str
    message: str


def parse_gaia_review_workbook(path: Path, queue: str) -> list[GaiaReviewRow]:
    """Parse a GAIA 11 一覧表 Excel export."""
    workbook = load_workbook(path, read_only=True, data_only=True)
    data_sheet = next(
        (sheet for sheet in workbook.worksheets if sheet.title != "設定"), None
    )
    if data_sheet is None:
        workbook.close()
        raise ValueError(f"GAIA一覧表のデータシートがありません: {path}")

    rows: list[GaiaReviewRow] = []
    for row_number, values in enumerate(
        data_sheet.iter_rows(min_row=7, max_col=68, values_only=True), start=7
    ):
        name = clean_text(values[1])
        specification = clean_text(values[16])
        unit = clean_text(values[31])
        message = clean_text(values[49])
        price = clean_text(values[67])
        if not name and not message:
            continue
        rows.append(
            GaiaReviewRow(
                queue=queue,
                workbook=path.name,
                gaia_row=row_number,
                name=name,
                specification=specification,
                unit=unit,
                price=price,
                message=message,
            )
        )
    workbook.close()
    return rows


def load_source_review(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def expected_source_fields(row: dict[str, str]) -> tuple[str, str, str]:
    name = row.get("gaia_item_name") or row.get("item_name") or ""
    specification = row.get("gaia_condition") or row.get("specification") or ""
    unit = row.get("unit") or ""
    if row.get("package_unit_status") == "EQUIVALENT" and row.get("package_unit"):
        unit = row["package_unit"]
    return name, specification, unit


def reconcile_rows(
    source_rows: list[dict[str, str]], review_rows: list[GaiaReviewRow]
) -> list[dict[str, str]]:
    matches: dict[int, list[GaiaReviewRow]] = {}
    orphans: list[GaiaReviewRow] = []
    used_by_queue: dict[str, set[int]] = {}

    for review in review_rows:
        review_name = normalize_match_text(review.name)
        review_spec = normalize_match_text(review.specification)
        review_unit = normalize_unit(review.unit)
        candidates: list[tuple[int, int]] = []
        for index, source in enumerate(source_rows):
            source_name, source_spec, source_unit = expected_source_fields(
                source
            )
            if normalize_match_text(source_name) != review_name:
                continue
            score = 10
            if normalize_match_text(source_spec) == review_spec:
                score += 40
            if normalize_unit(source_unit) == review_unit:
                score += 20
            candidates.append((score, index))

        if not candidates:
            orphans.append(review)
            continue

        used = used_by_queue.setdefault(review.queue, set())
        available = [candidate for candidate in candidates if candidate[1] not in used]
        pool = available or candidates
        _, source_index = max(pool, key=lambda candidate: (candidate[0], -candidate[1]))
        used.add(source_index)
        matches.setdefault(source_index, []).append(review)

    fieldnames = list(source_rows[0].keys()) if source_rows else []
    gaia_fields = [
        "gaia_review_status",
        "gaia_review_queue",
        "gaia_review_workbook",
        "gaia_review_row",
        "gaia_review_name",
        "gaia_review_specification",
        "gaia_review_unit",
        "gaia_review_price",
        "gaia_review_message",
    ]
    output: list[dict[str, str]] = []
    for index, source in enumerate(source_rows):
        reviews = matches.get(index, [])
        queues = sorted({review.queue for review in reviews})
        if queues == ["unit"]:
            review_status = "GAIA_UNIT_REVIEW"
        elif queues == ["work"]:
            review_status = "GAIA_WORK_REVIEW"
        elif queues:
            review_status = "GAIA_UNIT_AND_WORK_REVIEW"
        else:
            review_status = "NOT_IN_REVIEW_EXPORT"

        def joined(attribute: str, separator: str = " | ") -> str:
            values: list[str] = []
            for review in reviews:
                value = clean_text(getattr(review, attribute))
                if value and value not in values:
                    values.append(value)
            return separator.join(values)

        result = {name: source.get(name, "") for name in fieldnames}
        result.update(
            {
                "gaia_review_status": review_status,
                "gaia_review_queue": " | ".join(queues),
                "gaia_review_workbook": joined("workbook"),
                "gaia_review_row": " | ".join(
                    str(review.gaia_row) for review in reviews
                ),
                "gaia_review_name": joined("name"),
                "gaia_review_specification": joined("specification"),
                "gaia_review_unit": joined("unit"),
                "gaia_review_price": joined("price"),
                "gaia_review_message": joined("message"),
            }
        )
        output.append(result)

    for review in orphans:
        result = {name: "" for name in fieldnames}
        result.update(
            {
                "gaia_review_status": "UNMATCHED_GAIA_REVIEW_ROW",
                "gaia_review_queue": review.queue,
                "gaia_review_workbook": review.workbook,
                "gaia_review_row": str(review.gaia_row),
                "gaia_review_name": review.name,
                "gaia_review_specification": review.specification,
                "gaia_review_unit": review.unit,
                "gaia_review_price": review.price,
                "gaia_review_message": review.message,
            }
        )
        output.append(result)
    return output


def write_reconciliation(path: Path, rows: list[dict[str, str]]) -> dict[str, int]:
    path.parent.mkdir(parents=True, exist_ok=True)
    if not rows:
        raise ValueError("照合対象の行がありません")
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)

    counts: dict[str, int] = {}
    for row in rows:
        status = row["gaia_review_status"]
        counts[status] = counts.get(status, 0) + 1
    summary_path = path.with_suffix(".json")
    summary_path.write_text(
        json.dumps(
            {"output": str(path.resolve()), "status_counts": counts},
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )
    return counts


def main() -> int:
    parser = argparse.ArgumentParser(
        description="GAIAの確認工種・確認単価Excelを変換レビューCSVと照合します。"
    )
    parser.add_argument("--source-review", required=True, type=Path)
    parser.add_argument("--unit-review", type=Path)
    parser.add_argument("--work-review", type=Path)
    parser.add_argument("--output", required=True, type=Path)
    args = parser.parse_args()
    if not args.unit_review and not args.work_review:
        parser.error("--unit-review または --work-review が必要です")

    review_rows: list[GaiaReviewRow] = []
    if args.unit_review:
        review_rows.extend(parse_gaia_review_workbook(args.unit_review, "unit"))
    if args.work_review:
        review_rows.extend(parse_gaia_review_workbook(args.work_review, "work"))
    reconciled = reconcile_rows(load_source_review(args.source_review), review_rows)
    counts = write_reconciliation(args.output, reconciled)
    print(
        json.dumps(
            {
                "gaia_review_rows": len(review_rows),
                "reconciled_rows": len(reconciled),
                "status_counts": counts,
                "output": str(args.output.resolve()),
            },
            ensure_ascii=False,
            indent=2,
        )
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
