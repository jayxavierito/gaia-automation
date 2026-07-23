from __future__ import annotations

import hashlib
import json
import re
import unicodedata
from dataclasses import asdict, dataclass
from difflib import SequenceMatcher
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook


GAIA_CODE_RE = re.compile(r"^\[([A-Z]{2}\d{6})\]$")
HIERARCHY_MARKERS = {
    "費目行": ("level1", 0),
    "工種行": ("level2", 1),
    "種別行": ("level3", 2),
}
MATCH_PUNCTUATION_RE = re.compile(
    r"[\s・･()（）「」【】\[\]{}｛｝/／,，.。:：;；"
    r"‐‑‒–—―ーｰ\-_'\"“”‘’]"
)


def clean_text(value: object) -> str:
    if value is None:
        return ""
    text = unicodedata.normalize("NFKC", str(value))
    return re.sub(r"\s+", " ", text).strip()


def normalize_name(value: object) -> str:
    text = clean_text(value)
    text = text.replace("コンクリ-ト", "コンクリート")
    text = text.replace("積込み", "積込")
    normalized = MATCH_PUNCTUATION_RE.sub("", text).lower()
    return normalized[:-1] if len(normalized) > 2 and normalized.endswith("工") else normalized


def normalize_unit(value: object) -> str:
    text = clean_text(value).lower()
    replacements = {
        "m²": "m2",
        "m³": "m3",
        "㎡": "m2",
        "㎥": "m3",
        "ton": "t",
    }
    for old, new in replacements.items():
        text = text.replace(old, new)
    return text.replace(" ", "")


def units_compatible(left: object, right: object) -> bool:
    left_unit = normalize_unit(left)
    right_unit = normalize_unit(right)
    if not left_unit or not right_unit:
        return True
    if left_unit == right_unit:
        return True
    return (left_unit, right_unit) in {
        ("孔", "箇所"),
        ("箇所", "孔"),
        ("掛m2", "m2"),
        ("m2", "掛m2"),
    }


def text_similarity(left: object, right: object) -> float:
    left_key = normalize_name(left)
    right_key = normalize_name(right)
    if not left_key or not right_key:
        return 0.0
    if left_key == right_key:
        return 1.0
    shorter, longer = sorted((left_key, right_key), key=len)
    if len(shorter) >= 3 and shorter in longer:
        return max(0.82, 0.98 * len(shorter) / len(longer))
    return SequenceMatcher(None, left_key, right_key).ratio()


def specification_similarity(left: object, right: object) -> float:
    left_key = normalize_name(left)
    right_key = normalize_name(right)
    if not left_key and not right_key:
        return 1.0
    if not left_key or not right_key:
        return 0.25
    return text_similarity(left_key, right_key)


@dataclass(frozen=True)
class GaiaCatalogEntry:
    level1: str
    level2: str
    level3: str
    item_name: str
    specification: str
    unit: str
    code: str
    source_file: str
    source_row: int

    @classmethod
    def from_dict(cls, value: dict) -> "GaiaCatalogEntry":
        return cls(
            level1=clean_text(value.get("level1")),
            level2=clean_text(value.get("level2")),
            level3=clean_text(value.get("level3")),
            item_name=clean_text(value.get("item_name")),
            specification=clean_text(value.get("specification")),
            unit=clean_text(value.get("unit")),
            code=clean_text(value.get("code")),
            source_file=clean_text(value.get("source_file")),
            source_row=int(value.get("source_row") or 0),
        )


@dataclass(frozen=True)
class GaiaCatalogMatch:
    status: str
    score: float
    name_score: float
    specification_score: float
    candidate_count: int
    path_safe: bool
    code_safe: bool
    entry: GaiaCatalogEntry | None = None


@dataclass(frozen=True)
class _ScoredEntry:
    entry: GaiaCatalogEntry
    score: float
    name_score: float
    specification_score: float
    context_score: float


def extract_gaia_catalog(source_path: Path) -> list[GaiaCatalogEntry]:
    source_path = Path(source_path)
    workbook = load_workbook(source_path, data_only=False, read_only=True)
    try:
        if "本工事費内訳表" not in workbook.sheetnames:
            raise ValueError(
                f"本工事費内訳表がないためGAIAカタログを作成できません: {source_path}"
            )
        sheet = workbook["本工事費内訳表"]
        rows = list(sheet.iter_rows(min_col=1, max_col=13, values_only=True))
    finally:
        workbook.close()

    current = {"level1": "", "level2": "", "level3": ""}
    entries: list[GaiaCatalogEntry] = []
    for index, values in enumerate(rows):
        marker = clean_text(values[12])
        marker_definition = HIERARCHY_MARKERS.get(marker)
        if marker_definition and index > 0:
            role, column = marker_definition
            name = clean_text(rows[index - 1][column])
            if not name:
                continue
            current[role] = name
            if role == "level1":
                current["level2"] = ""
                current["level3"] = ""
            elif role == "level2":
                current["level3"] = ""
            continue

        item_name = clean_text(values[3])
        unit = clean_text(values[9])
        quantity = values[7]
        if (
            not item_name
            or not unit
            or isinstance(quantity, bool)
            or not isinstance(quantity, (int, float))
        ):
            continue
        next_values = rows[index + 1] if index + 1 < len(rows) else ()
        specification = clean_text(next_values[3]) if next_values else ""
        code_match = GAIA_CODE_RE.fullmatch(marker)
        entries.append(
            GaiaCatalogEntry(
                level1=current["level1"],
                level2=current["level2"],
                level3=current["level3"],
                item_name=item_name,
                specification=specification,
                unit=unit,
                code=code_match.group(1) if code_match else "",
                source_file=source_path.name,
                source_row=index + 1,
            )
        )
    return entries


def write_gaia_catalog(
    source_paths: Iterable[Path], output_path: Path
) -> list[GaiaCatalogEntry]:
    sources: list[dict] = []
    entries: list[GaiaCatalogEntry] = []
    seen: set[tuple[str, ...]] = set()
    for source_path in source_paths:
        source_path = Path(source_path)
        source_entries = extract_gaia_catalog(source_path)
        digest = hashlib.sha256(source_path.read_bytes()).hexdigest()
        sources.append(
            {
                "file_name": source_path.name,
                "sha256": digest,
                "entries": len(source_entries),
            }
        )
        for entry in source_entries:
            key = (
                entry.level1,
                entry.level2,
                entry.level3,
                entry.item_name,
                entry.specification,
                entry.unit,
                entry.code,
            )
            if key in seen:
                continue
            seen.add(key)
            entries.append(entry)

    payload = {
        "schema_version": 1,
        "sources": sources,
        "entry_count": len(entries),
        "entries": [asdict(entry) for entry in entries],
    }
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    return entries


def load_gaia_catalog(path: Path | None) -> list[GaiaCatalogEntry]:
    if path is None:
        return []
    path = Path(path)
    if not path.is_file():
        raise FileNotFoundError(f"GAIA取込カタログがありません: {path}")
    payload = json.loads(path.read_text(encoding="utf-8"))
    if payload.get("schema_version") != 1:
        raise ValueError("未対応のGAIA取込カタログです")
    return [
        GaiaCatalogEntry.from_dict(value)
        for value in payload.get("entries", [])
    ]


def _score_entry(
    entry: GaiaCatalogEntry,
    *,
    item_name: str,
    specification: str,
    unit: str,
    work_type: str,
    category: str,
) -> _ScoredEntry:
    name_score = text_similarity(item_name, entry.item_name)
    unit_score = (
        1.0
        if normalize_unit(unit) == normalize_unit(entry.unit)
        else 0.7
        if units_compatible(unit, entry.unit)
        else 0.0
    )
    specification_score = specification_similarity(
        specification, entry.specification
    )
    level2_score = text_similarity(work_type, entry.level2)
    level3_score = text_similarity(category, entry.level3)
    context_score = level2_score * 0.4 + level3_score * 0.6
    score = (
        name_score * 0.6
        + unit_score * 0.12
        + specification_score * 0.18
        + context_score * 0.1
    )
    return _ScoredEntry(
        entry=entry,
        score=score,
        name_score=name_score,
        specification_score=specification_score,
        context_score=context_score,
    )


def match_gaia_catalog(
    entries: Iterable[GaiaCatalogEntry],
    *,
    item_name: str,
    specification: str,
    unit: str,
    work_type: str,
    category: str,
) -> GaiaCatalogMatch:
    entry_list = list(entries)
    source_key = normalize_name(item_name)
    exact_entries = [
        entry for entry in entry_list if normalize_name(entry.item_name) == source_key
    ]
    pool = exact_entries
    exact_name = bool(pool)
    if not pool:
        named = [
            (text_similarity(item_name, entry.item_name), entry)
            for entry in entry_list
        ]
        best_name_score = max((score for score, _ in named), default=0.0)
        if best_name_score < 0.82:
            return GaiaCatalogMatch(
                status="NOT_FOUND",
                score=best_name_score,
                name_score=best_name_score,
                specification_score=0.0,
                candidate_count=0,
                path_safe=False,
                code_safe=False,
            )
        pool = [
            entry
            for score, entry in named
            if score >= max(0.82, best_name_score - 0.04)
        ]

    compatible = [
        entry for entry in pool if units_compatible(unit, entry.unit)
    ]
    if compatible:
        pool = compatible

    scored = sorted(
        (
            _score_entry(
                entry,
                item_name=item_name,
                specification=specification,
                unit=unit,
                work_type=work_type,
                category=category,
            )
            for entry in pool
        ),
        key=lambda candidate: candidate.score,
        reverse=True,
    )
    if not scored:
        return GaiaCatalogMatch(
            status="NOT_FOUND",
            score=0.0,
            name_score=0.0,
            specification_score=0.0,
            candidate_count=0,
            path_safe=False,
            code_safe=False,
        )

    top = scored[0]
    runner_up_score = scored[1].score if len(scored) > 1 else 0.0
    margin = top.score - runner_up_score
    paths = {
        (candidate.entry.level2, candidate.entry.level3)
        for candidate in scored
    }
    context_present = bool(normalize_name(work_type) or normalize_name(category))
    path_safe = bool(top.entry.level2 and top.entry.level3) and (
        len(paths) == 1
        or (
            margin >= 0.08
            and (
                top.context_score >= 0.72
                or top.specification_score >= 0.82
            )
        )
    ) and (not context_present or top.context_score >= 0.55)

    codes = {candidate.entry.code for candidate in scored if candidate.entry.code}
    specification_evidence = bool(normalize_name(specification))
    code_safe = exact_name and bool(top.entry.code) and (
        len(codes) == 1
        or (
            path_safe
            and margin >= 0.12
            and top.specification_score >= 0.78
        )
    ) and (
        (
            specification_evidence
            and top.specification_score >= 0.75
        )
        or top.context_score >= 0.75
    )

    if exact_name:
        status = (
            "EXACT_CODE"
            if code_safe
            else "EXACT_PATH"
            if path_safe
            else "EXACT_AMBIGUOUS"
        )
    elif (
        top.name_score >= 0.92
        and margin >= 0.08
        and path_safe
    ):
        status = "FUZZY_PATH"
        code_safe = False
    else:
        status = "AMBIGUOUS"
        path_safe = False
        code_safe = False

    return GaiaCatalogMatch(
        status=status,
        score=round(top.score, 3),
        name_score=round(top.name_score, 3),
        specification_score=round(top.specification_score, 3),
        candidate_count=len(scored),
        path_safe=path_safe,
        code_safe=code_safe,
        entry=top.entry,
    )
