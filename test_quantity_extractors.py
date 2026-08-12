import unittest
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook
from PIL import Image, ImageDraw

import quantity_extractors

from quantity_extractors import (
    _decode_repeated_ocr,
    _detect_table_grid,
    _extract_pdf_vector_tables,
    _pdf_header_columns,
    _select_pdf_summary_pages,
    _pdf_vector_table_schema,
    _records_from_pdf_rows,
    _records_from_pdf_vector_table,
    extract_excel_quantity_source,
    parse_quantity,
)


class QuantityExtractorTests(unittest.TestCase):
    def test_generic_excel_header_table_carries_hierarchy_and_skips_total(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "数量総括"
        sheet["B1"] = "新橋 数量総括表"
        headers = ["工種", "種別", "細別", "規格等", "単位", "数量", "備考"]
        for column, value in enumerate(headers, start=2):
            sheet.cell(3, column).value = value
        sheet.append([])
        values = ["補修工", "断面修復工", "左官工法", "モルタル", "m3", 1.2, ""]
        for column, value in enumerate(values, start=2):
            sheet.cell(4, column).value = value
        continuation = [0, 0, 0, "別規格", "m3", 0.3, ""]
        for column, value in enumerate(continuation, start=2):
            sheet.cell(5, column).value = value
        total = [0, 0, 0, "合計", "m3", 1.5, ""]
        for column, value in enumerate(total, start=2):
            sheet.cell(6, column).value = value

        path = Path.cwd() / "test_112_新橋_数量計算書.xlsx"
        try:
            workbook.save(path)
            result = extract_excel_quantity_source(path)
        finally:
            path.unlink(missing_ok=True)

        self.assertEqual(result.project_name, "新橋")
        self.assertEqual(len(result.records), 2)
        self.assertEqual(result.records[1].item_name, "左官工法")
        self.assertEqual(result.records[1].specification, "別規格")
        self.assertEqual(result.records[1].quantity, 0.3)

    def test_repeated_page_header_does_not_corrupt_hierarchy(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "設計数量総括表(下部工)"
        sheet["A1"] = "設計数量総括表"
        sheet["A2"] = "設計書名：珠洲市橋梁災害(舟山橋)"
        headers = ["工　種", "種　別", "細　別", "規　　格", "単位", "数　量", "摘　　要"]
        for column, value in enumerate(headers, start=1):
            sheet.cell(5, column).value = value
        # A work-type header row with no item on the same line, as in a real
        # 数量計算書 where the items live on the next printed page.
        sheet["A6"] = "橋台工"
        # Page break: the title/project-name/header block repeats mid-sheet.
        sheet["A9"] = "設計数量総括表"
        sheet["A10"] = "設計書名：珠洲市橋梁災害(舟山橋)"
        for column, value in enumerate(headers, start=1):
            sheet.cell(13, column).value = value
        sheet["B14"] = "作業土工"
        sheet["C15"] = "床掘り"
        sheet["D15"] = "土砂,小規模"
        sheet["E15"] = "m3"
        sheet["F15"] = 19.1

        path = Path.cwd() / "test_舟山橋_数量計算書.xlsx"
        try:
            workbook.save(path)
            result = extract_excel_quantity_source(path)
        finally:
            path.unlink(missing_ok=True)

        self.assertEqual(len(result.records), 1)
        record = result.records[0]
        self.assertEqual(record.work_type, "橋台工")
        self.assertEqual(record.category, "作業土工")
        self.assertEqual(record.item_name, "床掘り")

    def test_summary_section_prefix_does_not_replace_project_name(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "数量総括"
        sheet["B1"] = "§1.数量総括表　　　　　　仮谷橋　数量総括表"
        headers = ["工種", "種別", "細別", "規格等", "単位", "数量"]
        for column, value in enumerate(headers, start=2):
            sheet.cell(3, column).value = value
        values = ["1. 下部工補修工", "断面修復工", "左官工法", "", "m3", 0.1]
        for column, value in enumerate(values, start=2):
            sheet.cell(4, column).value = value

        path = Path.cwd() / "test_103_仮谷橋_数量計算書.xlsx"
        try:
            workbook.save(path)
            result = extract_excel_quantity_source(path)
        finally:
            path.unlink(missing_ok=True)

        self.assertEqual(result.project_name, "仮谷橋")

    def test_quantity_rounds_excel_binary_noise(self):
        self.assertEqual(parse_quantity(15.379999999999999), 15.38)
        self.assertEqual(parse_quantity(0.7999999999999999), 0.8)

    def test_pdf_rows_use_standard_columns_and_require_review(self):
        rows = [
            ["工種", "種別", "細別", "規格等", "単位", "数量", "備考"],
            ["支承補修工", "支承取替工", "支承設置工", "固定", "基", "2", ""],
            ["", "", "", "可動", "基", "2", ""],
        ]
        columns = {
            "work_type": 0,
            "category": 1,
            "item_name": 2,
            "specification": 3,
            "unit": 4,
            "quantity": 5,
            "remarks": 6,
        }

        records = _records_from_pdf_rows(rows, 0, columns, 1)

        self.assertEqual(len(records), 2)
        self.assertEqual(records[1].item_name, "支承設置工")
        self.assertEqual(records[1].quantity, 2)
        self.assertEqual(records[1].extraction_status, "OCR_REVIEW_REQUIRED")

    def test_ocr_six_column_road_header_maps_saimoku_to_specification(self):
        rows = [
            ["工種", "種別", "細目", "単位", "数量", "摘要"],
            ["舗装工", "表層工", "再生密粒度アスコン(20F)", "m2", "78.6", ""],
            ["", "", "t=5cm", "", "", ""],
        ]

        header_row, columns, warnings = _pdf_header_columns(rows, 6)
        records = _records_from_pdf_rows(rows, header_row, columns, 3)

        self.assertEqual(
            columns,
            {
                "work_type": 0,
                "category": 1,
                "item_name": 1,
                "specification": 2,
                "unit": 3,
                "quantity": 4,
                "remarks": 5,
            },
        )
        self.assertTrue(warnings)
        self.assertEqual(len(records), 1)
        self.assertEqual(records[0].work_type, "舗装工")
        self.assertEqual(records[0].category, "表層工")
        self.assertEqual(records[0].item_name, "表層工")
        self.assertEqual(
            records[0].specification, "再生密粒度アスコン(20F) / t=5cm"
        )

    def test_ocr_partial_six_column_header_still_uses_road_mapping(self):
        rows = [
            ["工種", "", "細目", "単位", "数量", "摘要"],
            ["舗装工", "表層工", "t=5cm", "m2", "78.6", ""],
        ]

        header_row, columns, warnings = _pdf_header_columns(rows, 6)
        records = _records_from_pdf_rows(rows, header_row, columns, 3)

        self.assertEqual(columns["item_name"], 1)
        self.assertEqual(columns["category"], 1)
        self.assertEqual(columns["specification"], 2)
        self.assertTrue(warnings)
        self.assertEqual(records[0].item_name, "表層工")
        self.assertEqual(records[0].category, "表層工")
        self.assertEqual(records[0].specification, "t=5cm")

    def test_ocr_pdf_hierarchy_is_carried_to_continuation_page(self):
        columns = {
            "work_type": 0,
            "item_name": 1,
            "specification": 2,
            "unit": 3,
            "quantity": 4,
            "remarks": 5,
        }
        state = {}
        first = _records_from_pdf_rows(
            [
                ["工種", "種別", "細目", "単位", "数量", "摘要"],
                ["取壊し工", "As舗装取壊し", "t=5cm", "m2", "78.6", ""],
            ],
            0,
            columns,
            3,
            state=state,
        )
        second = _records_from_pdf_rows(
            [
                ["工種", "種別", "細目", "単位", "数量", "摘要"],
                ["", "As殻処理", "", "m3", "3.9", ""],
            ],
            0,
            columns,
            4,
            state=state,
        )

        self.assertEqual(first[0].work_type, "取壊し工")
        self.assertEqual(second[0].work_type, "取壊し工")

    def test_ocr_pdf_hierarchy_is_reset_across_page_gap(self):
        columns = {
            "work_type": 0,
            "item_name": 1,
            "specification": 2,
            "unit": 3,
            "quantity": 4,
            "remarks": 5,
        }
        state = {}
        _records_from_pdf_rows(
            [
                ["工種", "種別", "細目", "単位", "数量", "摘要"],
                ["取壊し工", "As舗装取壊し", "t=5cm", "m2", "78.6", ""],
            ],
            0,
            columns,
            3,
            state=state,
        )
        separate = _records_from_pdf_rows(
            [
                ["工種", "種別", "細目", "単位", "数量", "摘要"],
                ["", "掘削", "土砂", "m3", "3.9", ""],
            ],
            0,
            columns,
            12,
            state=state,
        )

        self.assertEqual(separate[0].work_type, "")

    def test_ocr_summary_title_anchors_page_selection(self):
        class FakePage:
            def __init__(self, text):
                self.text = text

            def extract_text(self):
                return self.text

        class FakeReader:
            pages = [FakePage("数量総括表"), FakePage("詳細計算")]

            def __init__(self, _path):
                pass

        grids = [
            ([0, 10, 20, 30, 40, 50, 60], list(range(11))),
            ([0, 10, 20, 30, 40, 50, 60], list(range(26))),
        ]
        with (
            patch.object(quantity_extractors, "PdfReader", FakeReader),
            patch.object(quantity_extractors, "_make_title_crops"),
            patch.object(quantity_extractors, "_run_windows_ocr", return_value={}),
            patch.object(quantity_extractors, "_detect_table_grid", side_effect=grids),
        ):
            selected, _, _, _ = _select_pdf_summary_pages(
                Path("summary.pdf"),
                [Path("page-1.png"), Path("page-2.png")],
                Path.cwd(),
            )

        self.assertEqual(selected, [1])

    def test_detects_ruled_summary_table(self):
        path = Path.cwd() / "test_quantity_table.png"
        try:
            image = Image.new("RGB", (800, 600), "white")
            draw = ImageDraw.Draw(image)
            vertical = [50 + index * 100 for index in range(8)]
            horizontal = [50 + index * 50 for index in range(10)]
            for x in vertical:
                draw.line((x, horizontal[0], x, horizontal[-1]), fill="black", width=2)
            for y in horizontal:
                draw.line((vertical[0], y, vertical[-1], y), fill="black", width=2)
            image.save(path)
            image.close()

            detected_vertical, detected_horizontal = _detect_table_grid(path)
        finally:
            path.unlink(missing_ok=True)

        self.assertEqual(len(detected_vertical), 8)
        self.assertEqual(len(detected_horizontal), 10)

    def test_repeated_ocr_collapses_isolated_value(self):
        result = {"text": "4040404040", "lines": [{"words": [{"text": "4040404040"}]}]}
        self.assertEqual(_decode_repeated_ocr(result), "40")

    def test_vector_pdf_prefers_total_quantity_column(self):
        table = [
            ["種別", None, "規格", "単位", "数量", None, None, "摘要"],
            [None, None, None, None, "A1橋台", "A2橋台", "合計", None],
            ["躯体工", None, "", "", "", "", "", ""],
            ["コンクリート", None, "σck=30N/mm2", "m3", "105.2", "100.4", "205.6", ""],
            ["鉄筋", "SD345", "D13", "kg", "383", "610", "993", ""],
            [None, None, "D16～D25", "kg", "4936", "5018", "9954", ""],
            [None, None, "合計", "kg", "5319", "5628", "10947", ""],
        ]

        schema = _pdf_vector_table_schema(table)
        self.assertIsNotNone(schema)
        records = _records_from_pdf_vector_table(
            table, schema, 10, "橋台工数量集計表", "橋台工"
        )

        self.assertEqual(
            [record.quantity for record in records], [205.6, 993, 9954]
        )
        self.assertEqual(records[0].work_type, "橋台工")
        self.assertEqual(records[0].category, "躯体工")
        self.assertEqual(records[1].specification, "SD345 / D13")
        self.assertEqual(records[2].specification, "SD345 / D16~D25")
        self.assertEqual(records[0].extraction_status, "PDF_TEXT_REVIEW_REQUIRED")

    def test_vector_pdf_preserves_location_columns_when_total_is_absent(self):
        table = [
            ["種別", "規格", "単位", "数量", None, "摘要"],
            [None, None, None, "A1", "A2", None],
            ["掘削長", "レキ質土", "m", "13.38", "15.12", "全回転式"],
        ]

        schema = _pdf_vector_table_schema(table)
        self.assertIsNotNone(schema)
        records = _records_from_pdf_vector_table(
            table, schema, 11, "場所打ち杭 数量集計表", "場所打ち杭"
        )

        self.assertEqual([record.quantity for record in records], [13.38, 15.12])
        self.assertIn("数量区分: A1", records[0].remarks)
        self.assertIn("数量区分: A2", records[1].remarks)

    def test_vector_bridge_scope_change_resets_category(self):
        first_table = [
            ["種別", "規格", "単位", "数量", "摘要"],
            ["躯体工", "", "", "", ""],
            ["コンクリート", "σck=30N/mm2", "m3", "12.4", ""],
        ]
        second_table = [
            ["種別", "規格", "単位", "数量", "摘要"],
            ["鉄筋", "SD345", "kg", "993", ""],
        ]
        first_schema = _pdf_vector_table_schema(first_table)
        second_schema = _pdf_vector_table_schema(second_table)
        self.assertIsNotNone(first_schema)
        self.assertIsNotNone(second_schema)
        state = {}

        first = _records_from_pdf_vector_table(
            first_table,
            first_schema,
            1,
            "橋台工数量集計表",
            "橋台工",
            state=state,
        )
        second = _records_from_pdf_vector_table(
            second_table,
            second_schema,
            2,
            "橋脚工数量集計表",
            "橋脚工",
            state=state,
        )

        self.assertEqual(first[0].category, "躯体工")
        self.assertEqual(second[0].work_type, "橋脚工")
        self.assertEqual(second[0].category, "")

    def test_vector_road_table_uses_type_as_item_and_saimoku_as_specification(self):
        table = [
            ["工種", "種別", "細目", "単位", "数量", "摘要"],
            ["土工", "掘削", "", "m3", "3.9", ""],
            ["", "床掘", "", "m3", "10.1", ""],
            ["舗装工", "表層工", "再生密粒度アスコン(20F)", "m2", "78.6", ""],
            ["", "", "t=5cm", "", "", ""],
            ["", "上層路盤工", "粒調砕石(M-30)\nt=10cm", "m2", "39.2", ""],
        ]

        schema = _pdf_vector_table_schema(table)
        self.assertIsNotNone(schema)
        self.assertEqual(schema.layout, "work_inline_specification")
        records = _records_from_pdf_vector_table(
            table, schema, 3, "数量総括表", ""
        )

        self.assertEqual(
            [record.item_name for record in records],
            ["掘削", "床掘", "表層工", "上層路盤工"],
        )
        self.assertEqual(
            [record.category for record in records],
            ["掘削", "床掘", "表層工", "上層路盤工"],
        )
        self.assertEqual(
            records[2].specification,
            "再生密粒度アスコン(20F) / t=5cm",
        )
        self.assertEqual(records[3].specification, "粒調砕石(M-30) t=10cm")

    def test_vector_dedicated_specification_column_keeps_existing_hierarchy(self):
        table = [
            ["工種", "種別", "細別", "規格", "単位", "数量", "摘要"],
            ["道路土工", "掘削工", "掘削", "土砂 / 小規模", "m3", "157.4", ""],
        ]

        schema = _pdf_vector_table_schema(table)
        self.assertIsNotNone(schema)
        self.assertEqual(schema.layout, "work_name")
        records = _records_from_pdf_vector_table(
            table, schema, 2, "数量総括表", ""
        )

        self.assertEqual(records[0].work_type, "道路土工")
        self.assertEqual(records[0].category, "掘削工")
        self.assertEqual(records[0].item_name, "掘削")
        self.assertEqual(records[0].specification, "土砂 / 小規模")

    def test_vector_pdf_ignores_contents_and_stops_at_detail_section(self):
        summary_table = [
            ["工種", "名称", "規格", "単位", "数量", "摘要"],
            ["主桁工", "コンクリート", "σck=30N/mm2", "m3", "12.4", ""],
        ]
        detail_table = [
            ["工種", "名称", "規格", "単位", "数量", "摘要"],
            ["誤抽出", "詳細計算", "", "m", "999", ""],
        ]

        class FakePage:
            def __init__(self, text, tables):
                self.text = text
                self.tables = tables

            def extract_text(self):
                return self.text

            def extract_tables(self):
                return self.tables

        class FakeDocument:
            pages = [
                FakePage("目次\n1. 本体工数量総括表", []),
                FakePage("1. 数量総括表", [summary_table]),
                FakePage("2. 上部工数量計算書", [detail_table]),
            ]

            def __enter__(self):
                return self

            def __exit__(self, *_args):
                return False

        class FakePdfPlumber:
            @staticmethod
            def open(_path):
                return FakeDocument()

        with patch.object(quantity_extractors, "pdfplumber", FakePdfPlumber):
            result = _extract_pdf_vector_tables(Path("磐若橋_数量計算書.pdf"))

        self.assertIsNotNone(result)
        self.assertEqual(result.metadata["selected_pages"], [2])
        self.assertEqual(len(result.records), 1)
        self.assertEqual(result.records[0].item_name, "コンクリート")

    def test_vector_pdf_resumes_at_later_summary_section(self):
        summary_table = [
            ["工種", "種別", "細目", "単位", "数量", "摘要"],
            ["土工", "掘削", "", "m3", "3.9", ""],
        ]
        later_summary = [
            ["工種", "種別", "細目", "単位", "数量", "摘要"],
            ["舗装工", "表層工", "t=5cm", "m2", "78.6", ""],
        ]
        detail_table = [
            ["工種", "名称", "規格", "単位", "数量", "摘要"],
            ["誤抽出", "詳細計算", "", "m", "999", ""],
        ]

        class FakePage:
            def __init__(self, text, tables):
                self.text = text
                self.tables = tables

            def extract_text(self):
                return self.text

            def extract_tables(self):
                return self.tables

        class FakeDocument:
            pages = [
                FakePage("数量総括表 市道1号", [summary_table]),
                FakePage("数量計算書 市道1号-1", [detail_table]),
                FakePage("数量総括表 市道1号-1", [later_summary]),
            ]

            def __enter__(self):
                return self

            def __exit__(self, *_args):
                return False

        class FakePdfPlumber:
            @staticmethod
            def open(_path):
                return FakeDocument()

        with patch.object(quantity_extractors, "pdfplumber", FakePdfPlumber):
            result = _extract_pdf_vector_tables(Path("道路_数量計算書.pdf"))

        self.assertIsNotNone(result)
        self.assertEqual(result.metadata["selected_pages"], [1, 3])
        self.assertEqual([record.quantity for record in result.records], [3.9, 78.6])
        self.assertEqual(result.records[1].work_type, "舗装工")

    def test_vector_pdf_carries_work_type_between_summary_pages(self):
        first_table = [
            ["工種", "種別", "細目", "単位", "数量", "摘要"],
            ["取壊し工", "As舗装取壊し", "t=5cm", "m2", "78.6", ""],
        ]
        second_table = [
            ["工種", "種別", "細目", "単位", "数量", "摘要"],
            ["", "As殻処理", "", "m3", "3.9", ""],
        ]

        class FakePage:
            def __init__(self, text, tables):
                self.text = text
                self.tables = tables

            def extract_text(self):
                return self.text

            def extract_tables(self):
                return self.tables

        class FakeDocument:
            pages = [
                FakePage("数量総括表-1", [first_table]),
                FakePage("続き", [second_table]),
            ]

            def __enter__(self):
                return self

            def __exit__(self, *_args):
                return False

        class FakePdfPlumber:
            @staticmethod
            def open(_path):
                return FakeDocument()

        with patch.object(quantity_extractors, "pdfplumber", FakePdfPlumber):
            result = _extract_pdf_vector_tables(Path("市道433号線_数量計算書.pdf"))

        self.assertIsNotNone(result)
        self.assertEqual(result.metadata["selected_pages"], [1, 2])
        self.assertEqual(result.records[1].work_type, "取壊し工")

    def test_vector_pdf_keeps_bridge_summary_without_work_type_column(self):
        bridge_table = [
            ["種別", None, "規格", "単位", "数量", None, None, "摘要"],
            [None, None, None, None, "A1橋台", "A2橋台", "合計", None],
            ["躯体工", None, "", "", "", "", "", ""],
            [
                "コンクリート",
                None,
                "σck=30N/mm2",
                "m3",
                "105.2",
                "100.4",
                "205.6",
                "",
            ],
        ]

        class FakePage:
            def extract_text(self):
                return "橋台工数量集計表"

            def extract_tables(self):
                return [bridge_table]

        class FakeDocument:
            pages = [FakePage()]

            def __enter__(self):
                return self

            def __exit__(self, *_args):
                return False

        class FakePdfPlumber:
            @staticmethod
            def open(_path):
                return FakeDocument()

        with patch.object(quantity_extractors, "pdfplumber", FakePdfPlumber):
            result = _extract_pdf_vector_tables(Path("橋梁_数量計算書.pdf"))

        self.assertIsNotNone(result)
        self.assertEqual(result.records[0].work_type, "橋台工")
        self.assertEqual(result.records[0].item_name, "コンクリート")
        self.assertEqual(result.records[0].quantity, 205.6)

    def test_vector_pdf_carries_heading_only_work_type_to_next_page(self):
        first_table = [
            ["工種", "種別", "細目", "単位", "数量", "摘要"],
            ["土工", "掘削", "", "m3", "3.9", ""],
            ["舗装工", "", "", "", "", ""],
        ]
        second_table = [
            ["工種", "種別", "細目", "単位", "数量", "摘要"],
            ["", "表層工", "t=5cm", "m2", "78.6", ""],
        ]

        class FakePage:
            def __init__(self, text, tables):
                self.text = text
                self.tables = tables

            def extract_text(self):
                return self.text

            def extract_tables(self):
                return self.tables

        class FakeDocument:
            pages = [
                FakePage("数量総括表-1", [first_table]),
                FakePage("数量総括表-2", [second_table]),
            ]

            def __enter__(self):
                return self

            def __exit__(self, *_args):
                return False

        class FakePdfPlumber:
            @staticmethod
            def open(_path):
                return FakeDocument()

        with patch.object(quantity_extractors, "pdfplumber", FakePdfPlumber):
            result = _extract_pdf_vector_tables(Path("道路_数量計算書.pdf"))

        self.assertIsNotNone(result)
        self.assertEqual(result.records[1].work_type, "舗装工")


if __name__ == "__main__":
    unittest.main()
