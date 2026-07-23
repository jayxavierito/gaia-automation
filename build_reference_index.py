from __future__ import annotations

import argparse
import hashlib
import json
import re
import unicodedata
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from pypdf import PdfReader

try:
    import pdfplumber
except ImportError:
    pdfplumber = None


PACKAGE_TITLE_RE = re.compile(r"No\.\s*(\d+)\s*【\s*(.*?)\s*】")
PACKAGE_UNIT_RE = re.compile(r"積算単位\s*[:：]\s*([^\s>＞]+)")
FISCAL_YEAR_RE = re.compile(r"(?:令和|R)\s*(\d+)\s*年度", re.IGNORECASE)
EFFECTIVE_DATE_RE = re.compile(r"令和\s*(\d+)\s*年\s*(\d+)\s*月\s*(\d+)\s*日")
TREE_CATEGORY_RE = re.compile(r"事業区分\s*[:：]\s*([^】\]\r\n]+)")


def clean_text(value: object) -> str:
    if value is None:
        return ""
    text = unicodedata.normalize("NFKC", str(value))
    return re.sub(r"\s+", " ", text).strip()


def search_text(value: object) -> str:
    text = clean_text(value)
    text = text.replace("コンクリ-ト", "コンクリート")
    text = text.replace("積込み", "積込")
    normalized = re.sub(
        r"[\s・･()（）「」【】\[\]‐‑‒–—―-]", "", text
    ).lower()
    return normalized[:-1] if normalized.endswith("工") else normalized


def normalize_unit(value: object) -> str:
    text = clean_text(value).lower()
    replacements = {
        "m²": "m2",
        "m³": "m3",
        "㎡": "m2",
        "㎥": "m3",
        "ton": "t",
    }
    for old, new in replacements.items():
        text = text.replace(old, new)
    return text.replace(" ", "")


def fiscal_year_from_text(text: str) -> int | None:
    match = FISCAL_YEAR_RE.search(clean_text(text))
    if not match:
        return None
    return 2018 + int(match.group(1))


def effective_date_from_text(text: str) -> str | None:
    match = EFFECTIVE_DATE_RE.search(clean_text(text))
    if not match:
        return None
    year = 2018 + int(match.group(1))
    return date(year, int(match.group(2)), int(match.group(3))).isoformat()


def file_record(path: Path) -> dict[str, Any]:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    stat = path.stat()
    return {
        "path": str(path.resolve()),
        "size": stat.st_size,
        "modified": datetime.fromtimestamp(stat.st_mtime, timezone.utc).isoformat(),
        "sha256": digest.hexdigest(),
    }


def read_pdf_pages(path: Path) -> tuple[list[dict[str, Any]], str]:
    reader = PdfReader(path)
    pages: list[dict[str, Any]] = []
    first_page_text = ""
    for page_number, page in enumerate(reader.pages, start=1):
        raw = clean_text((page.extract_text() or "").replace("\x00", " "))
        if page_number == 1:
            first_page_text = raw
        heading = ""
        lines = [clean_text(line) for line in raw.splitlines() if clean_text(line)]
        if lines:
            heading = lines[0][:160]
        pages.append(
            {
                "page": page_number,
                "heading": heading,
                "search_text": search_text(raw),
            }
        )
    return pages, first_page_text


def _tree_level_for_x(x0: float, page_width: float) -> int | None:
    ratio = x0 / page_width
    if 0.02 <= ratio < 0.135:
        return 1
    if ratio < 0.25:
        return 2
    if ratio < 0.365:
        return 3
    if ratio < 0.48:
        return 4
    if ratio < 0.63:
        return 5
    return None


def _tree_page_category(text: str) -> str:
    match = TREE_CATEGORY_RE.search(text)
    return clean_text(match.group(1)) if match else ""


def _balanced_tree_label(value: str) -> bool:
    pairs = (("(", ")"), ("【", "】"), ("[", "]"))
    for opening, closing in pairs:
        balance = 0
        for character in value:
            if character == opening:
                balance += 1
            elif character == closing:
                balance -= 1
            if balance < 0:
                return False
        if balance:
            return False
    return True


def _tree_page_events(page) -> list[tuple[float, int, str]]:
    grouped: list[dict[str, Any]] = []
    for word in page.extract_words(
        x_tolerance=1,
        y_tolerance=2,
        use_text_flow=False,
        keep_blank_chars=False,
    ):
        top = float(word["top"])
        if top < 75 or top > float(page.height) - 8:
            continue
        level = _tree_level_for_x(float(word["x0"]), float(page.width))
        if level is None:
            continue
        text = clean_text(word["text"])
        if not text:
            continue
        existing = next(
            (
                entry
                for entry in grouped
                if entry["level"] == level and abs(entry["top"] - top) <= 1.5
            ),
            None,
        )
        if existing is None:
            grouped.append({"top": top, "level": level, "parts": [text]})
        else:
            existing["parts"].append(text)

    events = [
        (entry["top"], entry["level"], "".join(entry["parts"]))
        for entry in grouped
    ]
    events.sort(key=lambda event: (event[0], event[1]))

    merged: list[tuple[float, int, str]] = []
    for top, level, text in events:
        previous_has_open_pair = bool(
            merged
            and (
                merged[-1][2].count("(") > merged[-1][2].count(")")
                or merged[-1][2].count("【") > merged[-1][2].count("】")
            )
        )
        if (
            merged
            and level == merged[-1][1]
            and top - merged[-1][0] <= 10
            and (text in {"工", "等", "費", "部"} or previous_has_open_pair)
        ):
            previous_top, previous_level, previous_text = merged[-1]
            merged[-1] = (
                previous_top,
                previous_level,
                previous_text + text,
            )
            continue
        merged.append((top, level, text))
    return merged


def parse_level_tree_paths(path: Path) -> list[dict[str, Any]]:
    if pdfplumber is None:
        return []

    current_category = ""
    current_levels = {level: "" for level in range(1, 6)}
    by_path: dict[tuple[str, str, str, str, str], dict[str, Any]] = {}
    with pdfplumber.open(path) as pdf:
        for page_number, page in enumerate(pdf.pages, start=1):
            page_text = page.extract_text() or ""
            page_category = _tree_page_category(page_text)
            if page_category and page_category != current_category:
                current_category = page_category
                current_levels = {level: "" for level in range(1, 6)}

            for _, level, value in _tree_page_events(page):
                current_levels[level] = value
                for deeper_level in range(level + 1, 6):
                    current_levels[deeper_level] = ""
                if (
                    level != 4
                    or not current_category
                    or len(current_levels[4]) < 2
                ):
                    continue

                key = (
                    current_category,
                    current_levels[1],
                    current_levels[2],
                    current_levels[3],
                    current_levels[4],
                )
                if not all(key) or not all(
                    _balanced_tree_label(value) for value in key
                ):
                    continue
                if key not in by_path:
                    by_path[key] = {
                        "business_category": current_category,
                        "level1": current_levels[1],
                        "level2": current_levels[2],
                        "level3": current_levels[3],
                        "level4": current_levels[4],
                        "pages": [page_number],
                    }
                elif page_number not in by_path[key]["pages"]:
                    by_path[key]["pages"].append(page_number)
    return list(by_path.values())


def package_condition_fields(sheet) -> list[str]:
    condition_end = 0
    for merged in sheet.merged_cells.ranges:
        if not (merged.min_row <= 3 <= merged.max_row):
            continue
        anchor = clean_text(sheet.cell(merged.min_row, merged.min_col).value)
        if anchor == "条件区分":
            condition_end = merged.max_col
            break
    if not condition_end:
        for column in range(1, sheet.max_column + 1):
            if clean_text(sheet.cell(3, column).value) == "標準単価":
                condition_end = column - 1
                break
    fields: list[str] = []
    for column in range(1, condition_end + 1):
        label = clean_text(sheet.cell(4, column).value)
        if label and label not in fields:
            fields.append(label)
    return fields


def package_unit(sheet) -> str:
    text = clean_text(sheet.cell(2, 1).value)
    match = PACKAGE_UNIT_RE.search(text)
    return normalize_unit(match.group(1)) if match else ""


def package_title(sheet) -> tuple[str, str] | None:
    match = PACKAGE_TITLE_RE.search(clean_text(sheet.cell(1, 1).value))
    if not match:
        return None
    return match.group(1).zfill(3), clean_text(match.group(2))


def workbook_list_metadata(workbook) -> dict[str, dict[str, str]]:
    if "一覧" not in workbook.sheetnames:
        return {}
    sheet = workbook["一覧"]
    result: dict[str, dict[str, str]] = {}
    for row in range(1, sheet.max_row + 1):
        raw_no = sheet.cell(row, 1).value
        if raw_no in (None, ""):
            continue
        number = clean_text(raw_no).zfill(3)
        if not re.fullmatch(r"\d{3}", number):
            continue
        result[number] = {
            "list_name": clean_text(sheet.cell(row, 2).value),
            "list_page": clean_text(sheet.cell(row, 3).value),
            "standard_part": clean_text(sheet.cell(row, 4).value),
            "standard_chapter": clean_text(sheet.cell(row, 5).value),
            "standard_item": clean_text(sheet.cell(row, 6).value),
            "standard_title": clean_text(sheet.cell(row, 7).value),
            "list_notes": clean_text(sheet.cell(row, 8).value),
        }
    return result


def parse_package_workbook(
    path: Path,
    *,
    source_kind: str,
    fiscal_year: int,
    effective_from: str,
    scope: str,
    priority: int,
) -> list[dict[str, Any]]:
    workbook = load_workbook(path, data_only=True, read_only=False)
    list_metadata = workbook_list_metadata(workbook)
    packages: list[dict[str, Any]] = []
    for sheet in workbook.worksheets:
        title = package_title(sheet)
        if title is None:
            continue
        package_no, name = title
        metadata = list_metadata.get(package_no, {})
        packages.append(
            {
                "package_no": package_no,
                "name": name,
                "search_name": search_text(name),
                "unit": package_unit(sheet),
                "condition_fields": package_condition_fields(sheet),
                "source_kind": source_kind,
                "source_file": str(path.resolve()),
                "source_page": metadata.get("list_page", ""),
                "fiscal_year": fiscal_year,
                "effective_from": effective_from,
                "scope": scope,
                "priority": priority,
                "condition_schema_year": fiscal_year,
                "standard_reference": " / ".join(
                    value
                    for value in [
                        metadata.get("standard_part", ""),
                        metadata.get("standard_chapter", ""),
                        metadata.get("standard_item", ""),
                        metadata.get("standard_title", ""),
                    ]
                    if value
                ),
            }
        )
    workbook.close()
    return packages


def parse_package_pdf(
    path: Path,
    *,
    source_kind: str,
    scope: str,
    priority: int,
) -> tuple[list[dict[str, Any]], int | None, str | None]:
    reader = PdfReader(path)
    first_text = clean_text(reader.pages[0].extract_text() or "") if reader.pages else ""
    fiscal_year = fiscal_year_from_text(first_text)
    effective_from = effective_date_from_text(first_text)
    by_key: dict[tuple[str, str], dict[str, Any]] = {}
    for page_number, page in enumerate(reader.pages, start=1):
        text = clean_text((page.extract_text() or "").replace("\x00", " "))
        unit_match = PACKAGE_UNIT_RE.search(text)
        unit = normalize_unit(unit_match.group(1)) if unit_match else ""
        for match in PACKAGE_TITLE_RE.finditer(text):
            package_no = match.group(1).zfill(3)
            name = clean_text(match.group(2))
            key = (package_no, search_text(name))
            if key not in by_key:
                by_key[key] = {
                    "package_no": package_no,
                    "name": name,
                    "search_name": search_text(name),
                    "unit": unit,
                    "condition_fields": [],
                    "source_kind": source_kind,
                    "source_file": str(path.resolve()),
                    "source_page": str(page_number),
                    "source_page_end": str(page_number),
                    "fiscal_year": fiscal_year,
                    "effective_from": effective_from,
                    "scope": scope,
                    "priority": priority,
                    "condition_schema_year": None,
                    "standard_reference": "",
                }
            else:
                by_key[key]["source_page_end"] = str(page_number)
                if unit and not by_key[key]["unit"]:
                    by_key[key]["unit"] = unit
    return list(by_key.values()), fiscal_year, effective_from


def enrich_current_packages(
    current_packages: list[dict[str, Any]],
    schema_packages: list[dict[str, Any]],
) -> None:
    schema_by_name = {entry["search_name"]: entry for entry in schema_packages}
    for entry in current_packages:
        schema = schema_by_name.get(entry["search_name"])
        if not schema:
            continue
        if not entry["unit"]:
            entry["unit"] = schema["unit"]
        entry["condition_fields"] = schema["condition_fields"]
        entry["condition_schema_year"] = schema["fiscal_year"]
        entry["standard_reference"] = schema["standard_reference"]


def build_index(args: argparse.Namespace) -> dict[str, Any]:
    level_pages, _ = read_pdf_pages(args.level_tree)
    level_paths = parse_level_tree_paths(args.level_tree)
    guideline_pages, guideline_first = read_pdf_pages(args.quantity_guideline)

    national_schema = parse_package_workbook(
        args.national_reference_xlsx,
        source_kind="NATIONAL_REFERENCE_WORKBOOK",
        fiscal_year=args.national_reference_year,
        effective_from=f"{args.national_reference_year}-04-01",
        scope="全国標準(参考年度)",
        priority=100,
    )
    national_current, national_year, national_effective = parse_package_pdf(
        args.national_package_pdf,
        source_kind="NATIONAL_PACKAGE_PDF",
        scope="全国標準",
        priority=200,
    )
    enrich_current_packages(national_current, national_schema)

    ishikawa_first = ""
    if args.ishikawa_package_pdf:
        reader = PdfReader(args.ishikawa_package_pdf)
        if reader.pages:
            ishikawa_first = clean_text(reader.pages[0].extract_text() or "")
    ishikawa_year = fiscal_year_from_text(ishikawa_first) or national_year or 2026
    ishikawa_effective = (
        effective_date_from_text(ishikawa_first)
        or national_effective
        or f"{ishikawa_year}-04-01"
    )
    ishikawa_current = parse_package_workbook(
        args.ishikawa_package_xlsx,
        source_kind="ISHIKAWA_DISASTER_PACKAGE",
        fiscal_year=ishikawa_year,
        effective_from=ishikawa_effective,
        scope="石川県中能登・奥能登地域(珠洲市を含む)",
        priority=300,
    )

    sources = {
        "level_tree": file_record(args.level_tree),
        "quantity_guideline": file_record(args.quantity_guideline),
        "national_reference_xlsx": file_record(args.national_reference_xlsx),
        "national_package_pdf": file_record(args.national_package_pdf),
        "ishikawa_package_xlsx": file_record(args.ishikawa_package_xlsx),
    }
    if args.ishikawa_package_pdf:
        sources["ishikawa_package_pdf"] = file_record(args.ishikawa_package_pdf)

    return {
        "schema_version": 1,
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "sources": sources,
        "level_tree": {
            "pages": level_pages,
            "page_count": len(level_pages),
            "paths": level_paths,
            "path_count": len(level_paths),
        },
        "quantity_guideline": {
            "pages": guideline_pages,
            "page_count": len(guideline_pages),
            "fiscal_year": fiscal_year_from_text(guideline_first),
        },
        "package_catalog": ishikawa_current + national_current + national_schema,
    }


def main() -> int:
    parser = argparse.ArgumentParser(
        description="珠洲市GAIA変換用の積算体系・数量要領・施工パッケージ索引を作成します。"
    )
    parser.add_argument("--level-tree", required=True, type=Path)
    parser.add_argument("--quantity-guideline", required=True, type=Path)
    parser.add_argument("--national-reference-xlsx", required=True, type=Path)
    parser.add_argument("--national-reference-year", type=int, default=2025)
    parser.add_argument("--national-package-pdf", required=True, type=Path)
    parser.add_argument("--ishikawa-package-xlsx", required=True, type=Path)
    parser.add_argument("--ishikawa-package-pdf", type=Path)
    parser.add_argument("--output", required=True, type=Path)
    parser.add_argument(
        "--compact-tree-output",
        type=Path,
        help="Portable版向けに積算体系ツリーのレベル1～4だけを保存します。",
    )
    args = parser.parse_args()

    index = build_index(args)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(
        json.dumps(index, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    if args.compact_tree_output:
        compact_index = {
            "schema_version": 1,
            "generated_at": index["generated_at"],
            "sources": {},
            "level_tree": {
                "pages": [],
                "page_count": index["level_tree"]["page_count"],
                "paths": index["level_tree"]["paths"],
                "path_count": index["level_tree"]["path_count"],
            },
            "quantity_guideline": {"pages": [], "page_count": 0},
            "package_catalog": [],
        }
        args.compact_tree_output.parent.mkdir(parents=True, exist_ok=True)
        args.compact_tree_output.write_text(
            json.dumps(compact_index, ensure_ascii=False, separators=(",", ":")),
            encoding="utf-8",
        )
    summary = {
        "output": str(args.output.resolve()),
        "level_tree_pages": index["level_tree"]["page_count"],
        "quantity_guideline_pages": index["quantity_guideline"]["page_count"],
        "package_entries": len(index["package_catalog"]),
        "package_sources": {
            kind: sum(
                1
                for entry in index["package_catalog"]
                if entry["source_kind"] == kind
            )
            for kind in sorted(
                {entry["source_kind"] for entry in index["package_catalog"]}
            )
        },
    }
    print(json.dumps(summary, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
