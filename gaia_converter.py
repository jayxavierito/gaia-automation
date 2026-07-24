from __future__ import annotations

import argparse
import csv
import json
import re
import unicodedata
from collections import defaultdict
from dataclasses import asdict, dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.worksheet.worksheet import Worksheet
from pypdf import PdfReader

from gaia_catalog import (
    GaiaCatalogEntry,
    load_gaia_catalog,
    match_gaia_catalog,
)
from quantity_extractors import ExtractionResult, extract_quantity_source


ESTIMATE_DATE_LABELS = (
    "積算年月",
    "単価適用年月",
    "歩掛単価適用年月",
    "歩掛り単価適用年月",
    "単価年月",
)
ERA_YEAR_MONTH_RE = re.compile(
    r"(?:令和|R)\s*(\d{1,2})\s*(?:年|[./-])?\s*(\d{1,2})\s*月?",
    re.IGNORECASE,
)
WESTERN_YEAR_MONTH_RE = re.compile(
    r"\b(20\d{2})\s*(?:年|[./-])\s*(\d{1,2})\s*月?"
)
HIERARCHY_NUMBER_PREFIX_RES = (
    re.compile(r"^\s*[（(]\s*\d+(?:[.-]\d+)*\s*[）)]\s*"),
    re.compile(r"^\s*(?:§\s*)?\d+(?:[.-]\d+)*[.．、:：)]\s*"),
    re.compile(r"^\s*\d+\s+(?=\S)"),
)


@dataclass
class BoqItem:
    source_sheet: str
    source_row: int
    section: str
    work_type: str
    category: str
    item_name: str
    specification: str
    unit: str
    quantity: float | int | None
    remarks: str
    standard: str
    daily_output: str
    setting: str
    notes: str
    source_reference: str
    source_format: str = "excel"
    extractor: str = ""
    source_page: int | None = None
    extraction_status: str = "READY"
    extraction_confidence: float = 1.0
    extraction_warnings: list[str] = field(default_factory=list)
    gaia_item_name: str = ""
    gaia_code: str = ""
    gaia_condition: str = ""
    match_status: str = ""
    confidence: float = 0.0
    matched_rule: str = ""
    output_level1: str = ""
    output_level2: str = ""
    output_level3: str = ""
    output_level4: str = ""
    catalog_status: str = ""
    catalog_score: float = 0.0
    catalog_name_score: float = 0.0
    catalog_specification_score: float = 0.0
    catalog_candidate_count: int = 0
    catalog_path_safe: bool = False
    catalog_code_safe: bool = False
    catalog_source_row: int | None = None
    warnings: list[str] = field(default_factory=list)


def clean_text(value: object) -> str:
    if value is None:
        return ""
    text = unicodedata.normalize("NFKC", str(value))
    return re.sub(r"\s+", " ", text).strip()


def normalize_hierarchy_label(value: object) -> str:
    text = clean_text(value)
    for pattern in HIERARCHY_NUMBER_PREFIX_RES:
        cleaned = pattern.sub("", text)
        if cleaned != text:
            return cleaned.strip()
    return text


def normalize_match_text(value: object) -> str:
    text = clean_text(value).replace("コンクリ-ト", "コンクリート")
    text = text.replace("積込み", "積込")
    normalized = re.sub(
        r"[\s・･()（）「」【】\[\]‐‑‒–—―-]", "", text
    ).lower()
    return normalized[:-1] if normalized.endswith("工") else normalized


def normalize_unit(value: object) -> str:
    text = clean_text(value).lower()
    replacements = {
        "m²": "m2",
        "m³": "m3",
        "㎡": "m2",
        "㎥": "m3",
        "ton": "t",
    }
    for old, new in replacements.items():
        text = text.replace(old, new)
    return text.replace(" ", "")


def _year_month_from_value(value: object) -> date | None:
    if isinstance(value, datetime):
        return date(value.year, value.month, 1)
    if isinstance(value, date):
        return date(value.year, value.month, 1)
    text = clean_text(value)
    if not text:
        return None
    match = ERA_YEAR_MONTH_RE.search(text)
    if match:
        return date(2018 + int(match.group(1)), int(match.group(2)), 1)
    match = WESTERN_YEAR_MONTH_RE.search(text)
    if match:
        return date(int(match.group(1)), int(match.group(2)), 1)
    return None


def detect_estimate_date(source_path: Path) -> tuple[date, str]:
    suffix = source_path.suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        workbook = load_workbook(
            source_path,
            data_only=True,
            read_only=True,
        )
        try:
            for sheet in workbook.worksheets:
                max_row = min(sheet.max_row, 150)
                max_column = min(sheet.max_column, 40)
                for row in sheet.iter_rows(
                    min_row=1,
                    max_row=max_row,
                    min_col=1,
                    max_col=max_column,
                    values_only=True,
                ):
                    for column, value in enumerate(row):
                        text = clean_text(value)
                        if not any(label in text for label in ESTIMATE_DATE_LABELS):
                            continue
                        for candidate in row[column : column + 5]:
                            detected = _year_month_from_value(candidate)
                            if detected:
                                return detected, "SOURCE"
        finally:
            workbook.close()
    elif suffix == ".pdf":
        reader = PdfReader(source_path)
        text = "\n".join(
            page.extract_text() or "" for page in reader.pages[:15]
        )
        for line in text.splitlines():
            if not any(label in line for label in ESTIMATE_DATE_LABELS):
                continue
            detected = _year_month_from_value(line)
            if detected:
                return detected, "SOURCE"
    return date.today(), "PC_DATE"


def append_warning(item: BoqItem, message: str) -> None:
    cleaned = clean_text(message)
    if cleaned and cleaned not in item.warnings:
        item.warnings.append(cleaned)


def combine_unique(*values: str) -> str:
    result: list[str] = []
    seen: set[str] = set()
    for value in values:
        cleaned = clean_text(value)
        if not cleaned or cleaned in {"-", "―"} or cleaned in seen:
            continue
        seen.add(cleaned)
        result.append(cleaned)
    return " / ".join(result)


def boq_items_from_extraction(extraction: ExtractionResult) -> list[BoqItem]:
    items: list[BoqItem] = []
    for record in extraction.records:
        item = BoqItem(
            source_sheet=record.source_sheet,
            source_row=record.source_row,
            section=normalize_hierarchy_label(record.section),
            work_type=normalize_hierarchy_label(record.work_type),
            category=normalize_hierarchy_label(record.category),
            item_name=record.item_name,
            specification=record.specification,
            unit=record.unit,
            quantity=record.quantity,
            remarks=record.remarks,
            standard=record.standard,
            daily_output=record.daily_output,
            setting=record.setting,
            notes=record.notes,
            source_reference=record.source_reference,
            source_format=record.source_format,
            extractor=record.extractor,
            source_page=record.source_page,
            extraction_status=record.extraction_status,
            extraction_confidence=record.extraction_confidence,
            extraction_warnings=list(record.extraction_warnings),
        )
        item.gaia_item_name = item.item_name
        item.gaia_condition = combine_unique(item.specification, item.setting)
        for warning in item.extraction_warnings:
            append_warning(item, warning)
        items.append(item)
    return items


def apply_gaia_catalog(
    items: Iterable[BoqItem], catalog: Iterable[GaiaCatalogEntry]
) -> None:
    item_list = list(items)
    catalog_entries = list(catalog)
    for item in item_list:
        item.output_level1 = normalize_hierarchy_label(item.section) or "本工事費"
        item.output_level2 = normalize_hierarchy_label(item.work_type) or "工種未確認"
        item.output_level3 = normalize_hierarchy_label(item.category) or "種別未確認"
        item.output_level4 = clean_text(item.item_name)
        item.gaia_item_name = item.output_level4
        item.gaia_condition = combine_unique(item.specification, item.setting)

        match = match_gaia_catalog(
            catalog_entries,
            item_name=item.item_name,
            specification=item.gaia_condition,
            unit=item.unit,
            work_type=item.work_type,
            category=item.category,
        )
        item.catalog_status = match.status
        item.catalog_score = match.score
        item.catalog_name_score = match.name_score
        item.catalog_specification_score = match.specification_score
        item.catalog_candidate_count = match.candidate_count
        item.catalog_path_safe = match.path_safe
        item.catalog_code_safe = match.code_safe
        if match.entry is None:
            continue

        item.catalog_source_row = match.entry.source_row
        if match.status in {"EXACT_CODE", "EXACT_PATH", "FUZZY_PATH"}:
            item.output_level4 = match.entry.item_name
            item.gaia_item_name = match.entry.item_name
        if match.path_safe and not clean_text(item.work_type):
            item.output_level2 = match.entry.level2 or item.output_level2
        if match.path_safe and not clean_text(item.category):
            item.output_level3 = match.entry.level3 or item.output_level3
        if (
            match.code_safe
            and item.extraction_status == "READY"
            and item.quantity is not None
            and item.quantity > 0
        ):
            item.gaia_code = match.entry.code
        item.matched_rule = (
            f"GAIA_CATALOG:{match.entry.source_file}:"
            f"本工事費内訳表!{match.entry.source_row}"
        )

    # A confirmed item path is useful evidence for unmatched siblings in the
    # same source 工種/種別 group, but never supplies a GAIA code.
    grouped: dict[tuple[str, str, str, str], list[BoqItem]] = defaultdict(list)
    for item in item_list:
        work_key = normalize_match_text(item.work_type)
        category_key = normalize_match_text(item.category)
        if not work_key and not category_key:
            continue
        grouped[
            (
                clean_text(item.source_sheet),
                normalize_match_text(item.section),
                work_key,
                category_key,
            )
        ].append(item)
    for group_items in grouped.values():
        confirmed_paths = {
            (item.output_level2, item.output_level3)
            for item in group_items
            if item.catalog_path_safe
        }
        if len(confirmed_paths) != 1:
            continue
        level2, level3 = next(iter(confirmed_paths))
        for item in group_items:
            if item.catalog_path_safe:
                continue
            if not clean_text(item.work_type):
                item.output_level2 = level2
            if not clean_text(item.category):
                item.output_level3 = level3
            if item.catalog_status in {"NOT_FOUND", "AMBIGUOUS", "EXACT_AMBIGUOUS"}:
                item.catalog_status = f"{item.catalog_status}_GROUP_PATH"


def classify_schema_item(item: BoqItem) -> None:
    if item.quantity is None or item.quantity <= 0:
        item.match_status = "INVALID_QUANTITY"
        item.confidence = 0.0
        append_warning(item, "数量を取得できない、または数量が0以下です")
    elif item.extraction_status != "READY":
        item.match_status = "SOURCE_REVIEW_REQUIRED"
        item.confidence = min(0.5, item.extraction_confidence)
        append_warning(item, "原本照合が完了するまでGAIAコードを確定しません")
    elif item.gaia_code and item.catalog_code_safe:
        item.match_status = "CATALOG_EXACT_CODE"
        item.confidence = 0.98
    elif item.catalog_status in {"EXACT_PATH", "EXACT_CODE"}:
        item.match_status = "GAIA_NAME_SEARCH"
        item.confidence = 0.9
    elif item.catalog_status == "FUZZY_PATH":
        item.match_status = "GAIA_NAME_SEARCH_REVIEW"
        item.confidence = 0.65
    elif item.catalog_status.endswith("_GROUP_PATH"):
        item.match_status = "GAIA_NAME_SEARCH_GROUP"
        item.confidence = 0.55
    else:
        item.match_status = "GAIA_NAME_SEARCH_REQUIRED"
        item.confidence = 0.35


def infer_work_category(project_name: str, items: Iterable[BoqItem]) -> str:
    scores = {
        "橋梁工事": 0,
        "下水道工事（２）": 0,
        "道路工事": 0,
    }
    terms = {
        "橋梁工事": ("橋", "橋梁", "支承", "伸縮装置", "地覆", "胸壁", "床版", "主桁"),
        "下水道工事（２）": ("下水", "管渠", "マンホール", "汚水", "雨水", "推進工"),
        "道路工事": ("道路", "舗装", "路盤", "区画線", "側溝"),
    }
    project_text = clean_text(project_name)
    for category, keywords in terms.items():
        if any(keyword in project_text for keyword in keywords):
            scores[category] += 3
    for item in items:
        context = combine_unique(
            item.work_type,
            item.category,
            item.item_name,
            item.specification,
        )
        for category, keywords in terms.items():
            if any(keyword in context for keyword in keywords):
                scores[category] += 1
    selected, score = max(scores.items(), key=lambda pair: pair[1])
    return selected if score >= 2 else ""


def extract_boq_items(
    source_path: Path, work_dir: Path | None = None
) -> tuple[str, list[BoqItem]]:
    extraction = extract_quantity_source(source_path, work_dir=work_dir)
    items = boq_items_from_extraction(extraction)
    return extraction.project_name, items


def clear_values(sheet: Worksheet) -> None:
    for row in sheet.iter_rows():
        for cell in row:
            if not isinstance(cell, MergedCell):
                cell.value = None


def set_cover_fields(
    workbook,
    project_name: str,
    location: str,
    work_category: str,
    district: str,
    price_date: date | None,
) -> None:
    sheet = workbook["鏡"]
    sheet["D6"] = project_name
    sheet["D7"] = location
    sheet["E8"] = None
    sheet["J20"] = None
    sheet["D21"] = work_category
    sheet["J21"] = price_date
    sheet["O20"] = district


@dataclass
class OutputBlock:
    level: str
    value: str
    marker: str = ""
    item: BoqItem | None = None


def build_output_blocks(items: Iterable[BoqItem]) -> list[OutputBlock]:
    blocks: list[OutputBlock] = []
    previous_section = ""
    previous_work_type = ""
    previous_category = ""

    for item in items:
        if item.match_status == "INVALID_QUANTITY":
            continue
        section = clean_text(item.output_level1 or item.section) or "本工事費"
        work_type = clean_text(item.output_level2 or item.work_type) or "工種未確認"
        category = clean_text(item.output_level3 or item.category) or "種別未確認"
        item_name = clean_text(
            item.output_level4 or item.gaia_item_name or item.item_name
        )
        if section != previous_section:
            blocks.append(OutputBlock("fee", section, "費目行"))
            previous_section = section
            previous_work_type = ""
            previous_category = ""
        if work_type != previous_work_type:
            blocks.append(OutputBlock("work", work_type, "工種行"))
            previous_work_type = work_type
            previous_category = ""
        if category != previous_category:
            blocks.append(OutputBlock("category", category, "種別行"))
            previous_category = category
        blocks.append(OutputBlock("item", item_name, item=item))
    return blocks


def write_page_header(sheet: Worksheet, page_index: int) -> None:
    start = 1 + page_index * 40
    sheet.cell(start, 1).value = "本工事費内訳表"
    sheet.cell(start + 1, 13).value = page_index + 2
    sheet.cell(start + 2, 1).value = "費目　工種　種別　細別・規格"
    sheet.cell(start + 2, 8).value = "数　　量"
    sheet.cell(start + 2, 10).value = "単　位"
    sheet.cell(start + 2, 11).value = "単　　価"
    sheet.cell(start + 2, 12).value = "金　　額"
    sheet.cell(start + 2, 13).value = "摘　　　　要"


def write_block(sheet: Worksheet, block_index: int, block: OutputBlock) -> None:
    page_index, slot = divmod(block_index, 9)
    row = 1 + page_index * 40 + 3 + slot * 4
    if slot == 0:
        write_page_header(sheet, page_index)

    if block.level != "item":
        column = {"fee": 1, "work": 2, "category": 3}[block.level]
        sheet.cell(row, column).value = block.value
        sheet.cell(row, 8).value = 1
        sheet.cell(row, 10).value = "式"
        sheet.cell(row + 1, 13).value = block.marker
        return

    item = block.item
    assert item is not None
    sheet.cell(row, 4).value = block.value
    sheet.cell(row, 8).value = item.quantity
    sheet.cell(row, 10).value = item.unit
    sheet.cell(row, 13).value = f"[{item.gaia_code}]" if item.gaia_code else ""
    sheet.cell(row + 1, 4).value = item.gaia_condition


def build_gaia_candidate(
    template_path: Path,
    output_path: Path,
    project_name: str,
    items: list[BoqItem],
    location: str,
    work_category: str,
    district: str,
    price_date: date | None,
) -> int:
    workbook = load_workbook(template_path, data_only=False, read_only=False)
    required = {
        "鏡",
        "本工事費内訳表",
        "内訳書",
        "明細書",
        "代価表",
        "単価表",
        "施工パッケージ",
    }
    missing = required.difference(workbook.sheetnames)
    if missing:
        raise ValueError(f"テンプレートに必要なシートがありません: {sorted(missing)}")

    set_cover_fields(
        workbook, project_name, location, work_category, district, price_date
    )
    main_sheet = workbook["本工事費内訳表"]
    clear_values(main_sheet)
    for sheet_name in required - {"鏡", "本工事費内訳表"}:
        clear_values(workbook[sheet_name])

    blocks = build_output_blocks(items)
    capacity = ((main_sheet.max_row - 1) // 40 + 1) * 9
    if len(blocks) > capacity:
        raise ValueError(
            f"テンプレート容量不足: {len(blocks)} blocks > {capacity} blocks"
        )
    for index, block in enumerate(blocks):
        write_block(main_sheet, index, block)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    workbook.close()
    return len(blocks)


def write_review_csv(path: Path, items: list[BoqItem]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    fieldnames = [
        "source_format",
        "extractor",
        "source_sheet",
        "source_page",
        "source_row",
        "extraction_status",
        "extraction_confidence",
        "extraction_warnings",
        "section",
        "work_type",
        "category",
        "item_name",
        "specification",
        "unit",
        "quantity",
        "standard",
        "daily_output",
        "setting",
        "remarks",
        "notes",
        "source_reference",
        "gaia_item_name",
        "gaia_code",
        "gaia_condition",
        "match_status",
        "confidence",
        "matched_rule",
        "output_level1",
        "output_level2",
        "output_level3",
        "output_level4",
        "catalog_status",
        "catalog_score",
        "catalog_name_score",
        "catalog_specification_score",
        "catalog_candidate_count",
        "catalog_path_safe",
        "catalog_code_safe",
        "catalog_source_row",
        "warnings",
    ]
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for item in items:
            row = asdict(item)
            row["extraction_warnings"] = "; ".join(item.extraction_warnings)
            row["warnings"] = "; ".join(item.warnings)
            writer.writerow({name: row.get(name, "") for name in fieldnames})


def parse_iso_date(value: str) -> date | None:
    return date.fromisoformat(value) if value else None


def main() -> int:
    parser = argparse.ArgumentParser(
        description="数量計算書の設計数量総括表をGaia取込候補へ変換します。"
    )
    parser.add_argument("--source", required=True, type=Path)
    parser.add_argument(
        "--extraction-work-dir",
        type=Path,
        help="PDFページとOCRセルを診断用に保存するディレクトリ。",
    )
    parser.add_argument("--template", required=True, type=Path)
    parser.add_argument(
        "--catalog",
        type=Path,
        default=Path(__file__).resolve().parent
        / "assets"
        / "gaia_import_catalog.json",
        help="GAIA取込実績から作成した名称・コードカタログ。",
    )
    parser.add_argument("--output", required=True, type=Path)
    parser.add_argument("--review", type=Path)
    parser.add_argument("--project-name", default="")
    parser.add_argument("--location", default="")
    parser.add_argument(
        "--work-category",
        default="",
        help="鏡シートの工種区分。省略時は工事名と明細から推定します。",
    )
    parser.add_argument("--district", default="珠洲")
    parser.add_argument(
        "--price-date",
        default="",
        help="YYYY-MM-DD。省略時は原本から検出し、無ければPC日付を使用します。",
    )
    args = parser.parse_args()

    project_name, items = extract_boq_items(
        args.source, work_dir=args.extraction_work_dir
    )
    if args.project_name:
        project_name = clean_text(args.project_name)
    if args.price_date:
        price_date = parse_iso_date(args.price_date)
        price_date_source = "USER"
    else:
        price_date, price_date_source = detect_estimate_date(args.source)
    catalog = load_gaia_catalog(args.catalog)
    apply_gaia_catalog(items, catalog)
    for item in items:
        classify_schema_item(item)
    work_category = clean_text(args.work_category) or infer_work_category(
        project_name, items
    )

    review_path = args.review or args.output.with_name(
        f"{args.output.stem}_review.csv"
    )
    write_review_csv(review_path, items)
    block_count = build_gaia_candidate(
        template_path=args.template,
        output_path=args.output,
        project_name=project_name,
        items=items,
        location=clean_text(args.location),
        work_category=work_category,
        district=clean_text(args.district),
        price_date=price_date,
    )

    counts: dict[str, int] = {}
    for item in items:
        counts[item.match_status] = counts.get(item.match_status, 0) + 1
    result = {
        "project_name": project_name,
        "source_items": len(items),
        "source_format": items[0].source_format if items else "",
        "extraction_review_required": sum(
            item.extraction_status != "READY" for item in items
        ),
        "output_blocks": block_count,
        "status_counts": counts,
        "catalog": str(args.catalog.resolve()),
        "catalog_entries": len(catalog),
        "catalog_name_matched": sum(
            item.catalog_status.startswith("EXACT")
            or item.catalog_status == "FUZZY_PATH"
            for item in items
        ),
        "catalog_codes_applied": sum(bool(item.gaia_code) for item in items),
        "price_date": price_date.isoformat(),
        "price_date_source": price_date_source,
        "work_category": work_category,
        "output": str(args.output.resolve()),
        "review": str(review_path.resolve()),
    }
    print(json.dumps(result, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
