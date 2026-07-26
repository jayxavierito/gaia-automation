from __future__ import annotations

import hashlib
import json
import math
import re
import shutil
import subprocess
import tempfile
import unicodedata
from dataclasses import asdict, dataclass, field
from difflib import SequenceMatcher
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from PIL import Image, ImageDraw, ImageFont, ImageOps
from pypdf import PdfReader

try:
    import pdfplumber
except ImportError:  # The OCR fallback remains usable in minimal installations.
    pdfplumber = None


LEGACY_SUMMARY_RE = re.compile(r"^設計数量総括表[（(](.+?)[）)]$")
PROJECT_LABEL_RE = re.compile(r"設計書名\s*[:：]\s*(.+)")
JAPANESE_CHARACTERS = r"一-龯々ぁ-ゖァ-ヺー"
STRUCTURAL_SUMMARY_PREFIX_RE = re.compile(
    r"^\s*§\s*\d+(?:[.-]\d+)*[.．]?\s*(?:設計)?数量総括表\s*"
)
STRUCTURAL_PROJECT_NAME_RE = re.compile(
    r"^(?:§\s*\d+(?:[.-]\d+)*[.．]?|(?:設計)?数量(?:総括表|計算書)|目次)$"
)


class QuantityExtractionError(RuntimeError):
    """Raised when a source cannot be converted without guessing its layout."""


@dataclass
class QuantityRecord:
    source_format: str
    extractor: str
    source_sheet: str
    source_page: int | None
    source_row: int
    section: str
    work_type: str
    category: str
    item_name: str
    specification: str
    unit: str
    quantity: float | int | None
    remarks: str = ""
    standard: str = ""
    daily_output: str = ""
    setting: str = ""
    notes: str = ""
    source_reference: str = ""
    extraction_status: str = "READY"
    extraction_confidence: float = 1.0
    extraction_warnings: list[str] = field(default_factory=list)


@dataclass
class ExtractionResult:
    source_path: str
    source_format: str
    extractor: str
    project_name: str
    records: list[QuantityRecord]
    document_warnings: list[str] = field(default_factory=list)
    metadata: dict[str, Any] = field(default_factory=dict)

    def to_dict(self) -> dict[str, Any]:
        result = asdict(self)
        result["record_count"] = len(self.records)
        return result


@dataclass(frozen=True)
class _ExcelTableCandidate:
    sheet_name: str
    header_row: int
    columns: dict[str, int]
    score: int
    summary_level: int


@dataclass(frozen=True)
class _PdfVectorTableSchema:
    header_row: int
    data_start_row: int
    layout: str
    work_type_column: int | None
    hierarchy_columns: tuple[int, ...]
    specification_column: int | None
    unit_column: int
    quantity_columns: tuple[int, ...]
    quantity_labels: tuple[str, ...]
    remarks_column: int | None


HEADER_ALIASES = {
    "section": {"費目", "費目名"},
    "work_type": {"工種", "工種名"},
    "category": {"種別", "種別名"},
    "item_name": {"細別", "細別名", "細目", "名称", "項目"},
    "specification": {"規格", "規格等", "形状寸法", "仕様"},
    "unit": {"単位"},
    "quantity": {"数量", "設計数量", "合計数量"},
    "remarks": {"備考", "摘要"},
    "standard": {"積算基準", "適用基準", "基準", "適用積算基準"},
    "daily_output": {"日当り施工量", "日施工量", "日当たり施工量"},
    "setting": {"設定", "施工条件", "条件", "設定内容"},
    "notes": {"注記", "注意事項", "特記事項"},
    "source_reference": {"参照", "参照頁", "根拠頁", "出典"},
}


def clean_text(value: object) -> str:
    if value is None:
        return ""
    text = unicodedata.normalize("NFKC", str(value))
    return re.sub(r"\s+", " ", text).strip()


def clean_label(value: object) -> str:
    text = clean_text(value)
    return re.sub(
        rf"(?<=[{JAPANESE_CHARACTERS}])\s+(?=[{JAPANESE_CHARACTERS}])",
        "",
        text,
    )


def parse_quantity(value: object) -> float | int | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        number = float(value)
    else:
        text = clean_text(value).replace(",", "").replace(" ", "")
        text = text.replace("．", ".").replace("−", "-").replace("ー", "-")
        match = re.fullmatch(r"[-+]?\d+(?:\.\d+)?", text)
        if not match:
            return None
        number = float(text)
    if not math.isfinite(number):
        return None
    # Formula caches can expose binary tails such as 15.379999999999999.
    number = round(number, 12)
    return int(number) if number.is_integer() else number


def _normalize_header(value: object) -> str:
    text = clean_text(value).lower()
    return re.sub(r"[\s・･()（）\[\]【】:：/／]", "", text)


NORMALIZED_HEADER_ALIASES = {
    role: {_normalize_header(alias) for alias in aliases}
    for role, aliases in HEADER_ALIASES.items()
}


def _header_role(value: object) -> str:
    normalized = _normalize_header(value)
    if not normalized:
        return ""
    for role, aliases in NORMALIZED_HEADER_ALIASES.items():
        if normalized in aliases:
            return role
    return ""


def _text_cell(value: object) -> str:
    # Formula-only layout cells are commonly cached as numeric zero in quantity books.
    if isinstance(value, (int, float)) and not isinstance(value, bool) and value == 0:
        return ""
    return clean_label(value)


def _summary_level(sheet_name: str) -> int:
    normalized = clean_text(sheet_name).replace(" ", "")
    if LEGACY_SUMMARY_RE.match(normalized) or "数量総括" in normalized:
        return 3
    if "数量集計" in normalized:
        return 2
    if "数量表" in normalized:
        return 1
    return 0


def _detect_excel_tables(workbook) -> list[_ExcelTableCandidate]:
    candidates: list[_ExcelTableCandidate] = []
    for sheet in workbook.worksheets:
        if sheet.sheet_state != "visible":
            continue
        best: _ExcelTableCandidate | None = None
        max_row = min(sheet.max_row, 120)
        max_column = min(sheet.max_column, 40)
        for row in range(1, max_row + 1):
            columns: dict[str, int] = {}
            for column in range(1, max_column + 1):
                role = _header_role(sheet.cell(row, column).value)
                if role and role not in columns:
                    columns[role] = column
            required = {"item_name", "unit", "quantity"}
            hierarchy_count = len(
                {"section", "work_type", "category"}.intersection(columns)
            )
            if not required.issubset(columns) or hierarchy_count < 1:
                continue
            level = _summary_level(sheet.title)
            score = len(columns) * 10 + hierarchy_count * 5 + level * 30
            candidate = _ExcelTableCandidate(
                sheet_name=sheet.title,
                header_row=row,
                columns=columns,
                score=score,
                summary_level=level,
            )
            if best is None or candidate.score > best.score:
                best = candidate
        if best is not None:
            candidates.append(best)
    return candidates


def _select_excel_tables(
    candidates: list[_ExcelTableCandidate],
) -> tuple[list[_ExcelTableCandidate], list[str]]:
    if not candidates:
        raise QuantityExtractionError(
            "工種・細別・単位・数量の見出しを持つ表が見つかりません。"
        )
    strongest_summary = max(candidate.summary_level for candidate in candidates)
    if strongest_summary >= 3:
        return (
            [
                candidate
                for candidate in candidates
                if candidate.summary_level == strongest_summary
            ],
            [],
        )
    best = max(candidates, key=lambda candidate: candidate.score)
    warning = (
        "数量総括表という名前のシートが無いため、見出し構成が最も近い表を"
        f"選択しました: {best.sheet_name}"
    )
    return [best], [warning]


def _project_from_filename(path: Path) -> str:
    name = re.sub(r"[（(]?数量(?:計算書|総括表)[）)]?", "", path.stem)
    name = re.sub(r"^\d+[\s_\-－]+", "", name)
    return name.strip(" _-－") or "数量計算書変換工事"


def _project_from_summary_title(value: object) -> str:
    text = clean_label(value)
    if not text:
        return ""
    match = PROJECT_LABEL_RE.search(text)
    if match:
        candidate = clean_text(match.group(1))
    else:
        legacy = re.fullmatch(r"(?:設計)?数量総括表[（(](.+?)[）)]", text)
        if legacy:
            candidate = clean_text(legacy.group(1))
        else:
            if "数量総括表" not in text:
                return ""
            candidate = STRUCTURAL_SUMMARY_PREFIX_RE.sub("", text)
            candidate = re.sub(
                r"\s*(?:設計)?数量総括表\s*$", "", candidate
            ).strip()
    if not candidate or STRUCTURAL_PROJECT_NAME_RE.fullmatch(candidate):
        return ""
    return candidate


def _extract_excel_project_name(
    workbook, selected: list[_ExcelTableCandidate], source_path: Path
) -> str:
    for candidate in selected:
        sheet = workbook[candidate.sheet_name]
        for row in range(1, min(candidate.header_row, 15) + 1):
            for column in range(1, min(sheet.max_column, 15) + 1):
                project_name = _project_from_summary_title(
                    sheet.cell(row, column).value
                )
                if project_name:
                    return project_name
    return _project_from_filename(source_path)


def _section_from_sheet(sheet_name: str) -> str:
    match = LEGACY_SUMMARY_RE.match(clean_text(sheet_name))
    return clean_text(match.group(1)) if match else "本工事費"


def _is_total_row(*values: str) -> bool:
    total_labels = {"計", "小計", "合計", "総計", "合計数量"}
    return any(clean_text(value).replace(" ", "") in total_labels for value in values)


def _excel_value(sheet: Worksheet, row: int, columns: dict[str, int], role: str):
    column = columns.get(role)
    return sheet.cell(row, column).value if column else None


def _quantity_on_or_below(
    sheet: Worksheet, row: int, columns: dict[str, int]
) -> tuple[float | int | None, int]:
    quantity = parse_quantity(_excel_value(sheet, row, columns, "quantity"))
    if quantity is not None:
        return quantity, row
    for next_row in range(row + 1, min(row + 4, sheet.max_row) + 1):
        if any(
            _text_cell(_excel_value(sheet, next_row, columns, role))
            for role in ("section", "work_type", "category", "item_name", "unit")
        ):
            break
        quantity = parse_quantity(
            _excel_value(sheet, next_row, columns, "quantity")
        )
        if quantity is not None:
            return quantity, next_row
    return None, row


def _with_legacy_columns(
    candidate: _ExcelTableCandidate,
) -> dict[str, int]:
    columns = dict(candidate.columns)
    if not LEGACY_SUMMARY_RE.match(candidate.sheet_name):
        return columns
    fallback = {
        "work_type": 1,
        "category": 2,
        "item_name": 3,
        "specification": 4,
        "unit": 5,
        "quantity": 6,
        "remarks": 7,
        "standard": 8,
        "daily_output": 9,
        "setting": 10,
        "notes": 11,
    }
    for role, column in fallback.items():
        columns.setdefault(role, column)
    return columns


def _is_repeated_header_block_row(
    sheet: Worksheet, row: int, columns: dict[str, int]
) -> bool:
    """Detect page-break title/header rows repeated mid-sheet.

    Long summary sheets restate the sheet title, project name, and column
    headers every printed page (e.g. every 40 rows). Those restated rows sit
    in the same columns as real data and would otherwise be read as a new
    hierarchy value (e.g. work_type becoming the literal text "工種"),
    silently corrupting the running section/work_type/category state for
    every row that follows until the next explicit value.
    """
    title_cell = clean_text(sheet.cell(row, 1).value)
    if title_cell and ("総括表" in title_cell or "数量計算書" in title_cell):
        return True
    if title_cell and PROJECT_LABEL_RE.search(title_cell):
        return True

    checked = 0
    matched = 0
    for column in columns.values():
        value = sheet.cell(row, column).value
        if value is None or (isinstance(value, str) and not value.strip()):
            continue
        checked += 1
        if _header_role(value):
            matched += 1
    return checked >= 2 and matched / checked >= 0.6


def _extract_excel_table(
    sheet: Worksheet, candidate: _ExcelTableCandidate
) -> list[QuantityRecord]:
    columns = _with_legacy_columns(candidate)
    section = _section_from_sheet(sheet.title)
    current = {
        "section": section,
        "work_type": "",
        "category": "",
        "item_name": "",
        "unit": "",
    }
    records: list[QuantityRecord] = []
    consumed_quantity_rows: set[int] = set()

    for row in range(candidate.header_row + 1, sheet.max_row + 1):
        if row in consumed_quantity_rows:
            continue
        if _is_repeated_header_block_row(sheet, row, columns):
            continue
        explicit = {
            role: _text_cell(_excel_value(sheet, row, columns, role))
            for role in (
                "section",
                "work_type",
                "category",
                "item_name",
                "specification",
                "unit",
                "remarks",
                "standard",
                "daily_output",
                "setting",
                "notes",
                "source_reference",
            )
        }
        if _is_total_row(
            explicit["section"],
            explicit["work_type"],
            explicit["category"],
            explicit["item_name"],
            explicit["specification"],
        ):
            continue

        if explicit["section"]:
            current["section"] = explicit["section"]
            current["work_type"] = ""
            current["category"] = ""
            current["item_name"] = ""
        if explicit["work_type"]:
            current["work_type"] = explicit["work_type"]
            current["category"] = ""
            current["item_name"] = ""
        if explicit["category"]:
            current["category"] = explicit["category"]
            current["item_name"] = ""
        if explicit["item_name"]:
            current["item_name"] = explicit["item_name"]
        if explicit["unit"]:
            current["unit"] = explicit["unit"]

        quantity, quantity_row = _quantity_on_or_below(sheet, row, columns)
        if quantity_row != row:
            consumed_quantity_rows.add(quantity_row)
        has_row_evidence = any(explicit.values()) or quantity is not None
        if not has_row_evidence:
            continue

        item_name = explicit["item_name"] or current["item_name"]
        unit = explicit["unit"] or (
            current["unit"]
            if explicit["specification"] or quantity is not None
            else ""
        )
        if not item_name or not unit:
            continue

        # Ignore formula artefacts that inherit the previous label but contain no work.
        if (
            quantity == 0
            and not explicit["item_name"]
            and not explicit["specification"]
            and not explicit["remarks"]
        ):
            continue

        warnings: list[str] = []
        status = "READY"
        confidence = 0.98 if candidate.summary_level >= 3 else 0.9
        if quantity is None:
            status = "SOURCE_REVIEW_REQUIRED"
            confidence = min(confidence, 0.65)
            warnings.append("数量を読み取れませんでした。")

        source_reference = explicit["source_reference"]
        if quantity_row != row and columns.get("standard"):
            source_reference = source_reference or _text_cell(
                sheet.cell(quantity_row, columns["standard"]).value
            )

        records.append(
            QuantityRecord(
                source_format="excel",
                extractor="excel_header_table",
                source_sheet=sheet.title,
                source_page=None,
                source_row=row,
                section=current["section"] or section,
                work_type=current["work_type"],
                category=current["category"],
                item_name=item_name,
                specification=explicit["specification"],
                unit=unit,
                quantity=quantity,
                remarks=explicit["remarks"],
                standard=explicit["standard"],
                daily_output=explicit["daily_output"],
                setting=explicit["setting"],
                notes=explicit["notes"],
                source_reference=source_reference,
                extraction_status=status,
                extraction_confidence=confidence,
                extraction_warnings=warnings,
            )
        )
    return records


def extract_excel_quantity_source(source_path: Path) -> ExtractionResult:
    workbook = load_workbook(source_path, data_only=True, read_only=False)
    try:
        candidates = _detect_excel_tables(workbook)
        selected, warnings = _select_excel_tables(candidates)
        project_name = _extract_excel_project_name(workbook, selected, source_path)
        records: list[QuantityRecord] = []
        for candidate in selected:
            records.extend(
                _extract_excel_table(workbook[candidate.sheet_name], candidate)
            )
    finally:
        workbook.close()

    if not records:
        raise QuantityExtractionError(
            "数量総括表は見つかりましたが、単位付きの明細を抽出できませんでした。"
        )
    return ExtractionResult(
        source_path=str(source_path.resolve()),
        source_format="excel",
        extractor="excel_header_table",
        project_name=project_name,
        records=records,
        document_warnings=warnings,
        metadata={
            "selected_tables": [
                {
                    "sheet": candidate.sheet_name,
                    "header_row": candidate.header_row,
                    "columns": candidate.columns,
                }
                for candidate in selected
            ]
        },
    )


def _pdf_vector_cell(value: object) -> str:
    return clean_label(value)


def _pdf_vector_table_schema(
    table: list[list[object]],
) -> _PdfVectorTableSchema | None:
    for header_row, row in enumerate(table[:3]):
        roles: dict[str, int] = {}
        for column, value in enumerate(row):
            role = _header_role(value)
            if role and role not in roles:
                roles[role] = column

        if not {"unit", "quantity"}.issubset(roles):
            continue
        if not {"work_type", "category", "item_name"}.intersection(roles):
            continue

        specification_column = roles.get("specification")
        unit_column = roles["unit"]
        remarks_column = roles.get("remarks")
        quantity_start = roles["quantity"]
        quantity_end = remarks_column if remarks_column is not None else len(row)
        quantity_columns = tuple(range(quantity_start, quantity_end))
        if not quantity_columns:
            continue

        work_type_column = roles.get("work_type")
        if "item_name" in roles:
            layout = "work_name"
            hierarchy_start = roles.get("category", roles["item_name"])
        else:
            layout = "type"
            hierarchy_start = roles["category"]
        hierarchy_end = specification_column or unit_column
        hierarchy_columns = tuple(range(hierarchy_start, hierarchy_end))
        if not hierarchy_columns:
            continue

        data_start_row = header_row + 1
        quantity_labels = tuple("数量" for _ in quantity_columns)
        if data_start_row < len(table) and len(quantity_columns) > 1:
            possible_labels = tuple(
                _pdf_vector_cell(table[data_start_row][column])
                if column < len(table[data_start_row])
                else ""
                for column in quantity_columns
            )
            label_text = "".join(possible_labels)
            if label_text and re.search(rf"[A-Za-z{JAPANESE_CHARACTERS}]", label_text):
                quantity_labels = tuple(
                    label or f"数量{index + 1}"
                    for index, label in enumerate(possible_labels)
                )
                data_start_row += 1

        return _PdfVectorTableSchema(
            header_row=header_row,
            data_start_row=data_start_row,
            layout=layout,
            work_type_column=work_type_column,
            hierarchy_columns=hierarchy_columns,
            specification_column=specification_column,
            unit_column=unit_column,
            quantity_columns=quantity_columns,
            quantity_labels=quantity_labels,
            remarks_column=remarks_column,
        )
    return None


def _parse_pdf_vector_quantity(value: object) -> float | int | None:
    quantity = parse_quantity(value)
    if quantity is not None:
        return quantity
    text = clean_text(value).replace(",", "").replace("．", ".")
    if not text or re.fullmatch(r"[-―ー－–—]+", text):
        return None
    match = re.search(r"[-+]?\d+(?:\.\d+)?", text)
    return parse_quantity(match.group(0)) if match else None


def _pdf_vector_title_and_scope(page_text: str) -> tuple[str, str]:
    lines = [_pdf_vector_cell(line) for line in page_text.splitlines() if clean_text(line)]
    title = lines[0] if lines else ""
    scope = ""
    for line in lines[:8]:
        compact = re.sub(r"\s+", "", line)
        if "数量集計表" not in compact and "数量総括表" not in compact:
            continue
        title = line
        prefix = re.sub(r"(?:数量集計表|数量総括表).*$", "", compact)
        prefix = re.sub(r"^[第\d０-９]+[.．、・\s]*", "", prefix)
        if prefix:
            scope = prefix
        break
    return title, scope


def _is_pdf_summary_hint(page_text: str) -> bool:
    compact = re.sub(r"\s+", "", clean_text(page_text))
    return "数量総括表" in compact or "数量集計表" in compact


def _is_pdf_detail_section(page_text: str) -> bool:
    lines = [_pdf_vector_cell(line) for line in page_text.splitlines() if clean_text(line)]
    for line in lines[:6]:
        compact = re.sub(r"\s+", "", line)
        if (
            "数量計算書" in compact
            and "数量総括表" not in compact
            and "数量集計表" not in compact
        ):
            return True
    return False


def _pdf_vector_quantity_entries(
    row: list[object], schema: _PdfVectorTableSchema
) -> list[tuple[str, float | int, str, bool]]:
    parsed: list[tuple[str, float | int, str, bool]] = []
    total_column_exists = any(
        "合計" in re.sub(r"\s+", "", label) for label in schema.quantity_labels
    )
    for column, label in zip(schema.quantity_columns, schema.quantity_labels):
        raw_value = row[column] if column < len(row) else None
        quantity = _parse_pdf_vector_quantity(raw_value)
        if quantity is None:
            continue
        is_total = "合計" in re.sub(r"\s+", "", label)
        parsed.append((label, quantity, clean_text(raw_value), is_total))

    totals = [entry for entry in parsed if entry[3]]
    if totals:
        return totals[:1]
    if total_column_exists:
        return [entry for entry in parsed if not entry[3]]
    return parsed


def _append_pdf_note(base: str, note: str) -> str:
    return " / ".join(value for value in (clean_text(base), clean_text(note)) if value)


def _records_from_pdf_vector_table(
    table: list[list[object]],
    schema: _PdfVectorTableSchema,
    page_number: int,
    page_title: str,
    scope: str,
) -> list[QuantityRecord]:
    current_work_type = scope if schema.layout == "type" else ""
    current_category = ""
    current_item = ""
    current_type_qualifier = ""
    active_item_position = max(0, len(schema.hierarchy_columns) - 1)
    pending_item_names: list[str] = []
    records: list[QuantityRecord] = []

    for row_index in range(schema.data_start_row, len(table)):
        row = table[row_index]
        work_type = (
            _pdf_vector_cell(row[schema.work_type_column])
            if schema.work_type_column is not None
            and schema.work_type_column < len(row)
            else ""
        )
        hierarchy_raw = [
            row[column] if column < len(row) else None
            for column in schema.hierarchy_columns
        ]
        hierarchy = [_pdf_vector_cell(value) for value in hierarchy_raw]
        specification = (
            _pdf_vector_cell(row[schema.specification_column])
            if schema.specification_column is not None
            and schema.specification_column < len(row)
            else ""
        )
        unit = (
            _pdf_vector_cell(row[schema.unit_column])
            if schema.unit_column < len(row)
            else ""
        )
        remarks = (
            _pdf_vector_cell(row[schema.remarks_column])
            if schema.remarks_column is not None
            and schema.remarks_column < len(row)
            else ""
        )
        quantity_entries = _pdf_vector_quantity_entries(row, schema)

        if work_type:
            current_work_type = work_type
            current_category = ""
            current_item = ""
            pending_item_names = []
            if schema.layout == "work_name":
                nonempty_positions = [
                    index for index, value in enumerate(hierarchy) if value
                ]
                if nonempty_positions:
                    active_item_position = nonempty_positions[-1]

        extra_specification = ""
        item_cell_multiline = False
        item_cell_split = False
        if schema.layout == "work_name":
            category_parts = [
                value
                for index, value in enumerate(hierarchy)
                if index < active_item_position and value
            ]
            if category_parts:
                current_category = " / ".join(category_parts)

            item_value = (
                hierarchy[active_item_position] if hierarchy else ""
            )
            raw_item_value = (
                hierarchy_raw[active_item_position] if hierarchy_raw else None
            )
            raw_item_lines = [
                _pdf_vector_cell(line)
                for line in str(raw_item_value or "").splitlines()
                if _pdf_vector_cell(line)
            ]
            item_cell_multiline = len(raw_item_lines) > 1
            if len(raw_item_lines) == 2 and all(
                len(re.sub(r"\s+", "", line)) <= 6 for line in raw_item_lines
            ):
                next_row = table[row_index + 1] if row_index + 1 < len(table) else []
                next_item = (
                    next_row[schema.hierarchy_columns[active_item_position]]
                    if schema.hierarchy_columns
                    and schema.hierarchy_columns[active_item_position] < len(next_row)
                    else None
                )
                if (
                    not _pdf_vector_cell(next_item)
                    and _pdf_vector_quantity_entries(next_row, schema)
                    and not raw_item_lines[1].startswith("(")
                ):
                    item_value = raw_item_lines[0]
                    pending_item_names = raw_item_lines[1:]
                    item_cell_split = True
            if item_value:
                current_item = item_value
            elif pending_item_names:
                current_item = pending_item_names.pop(0)
        else:
            if not quantity_entries:
                heading = " / ".join(value for value in hierarchy if value)
                if heading and not _is_total_row(heading):
                    current_category = heading
                    current_item = ""
                    current_type_qualifier = ""
                continue
            if hierarchy and hierarchy[0]:
                current_item = hierarchy[0]
                current_type_qualifier = ""
            explicit_qualifier = " / ".join(
                value for value in hierarchy[1:] if value
            )
            if explicit_qualifier:
                current_type_qualifier = explicit_qualifier
            extra_specification = current_type_qualifier

        is_total_row = _is_total_row(*hierarchy, specification)
        item_name = current_item
        specification = _append_pdf_note(extra_specification, specification)
        if not item_name or not quantity_entries or not unit:
            continue
        if is_total_row or _is_total_row(item_name, specification):
            continue

        for label, quantity, raw_quantity, is_total in quantity_entries:
            row_remarks = remarks
            warnings = [
                "PDFの埋め込み文字と罫線から抽出しました。原本の単位・数量を確認してください。"
            ]
            is_grouped_quantity = len(schema.quantity_columns) > 1 and not is_total
            if is_grouped_quantity:
                row_remarks = _append_pdf_note(row_remarks, f"数量区分: {label}")
                warnings.append(
                    f"合計列が無いため、{label}列を独立した明細として保持しました。"
                )
            numeric_fragments = re.findall(r"[-+]?\d+(?:\.\d+)?", raw_quantity)
            if len(numeric_fragments) > 1:
                row_remarks = _append_pdf_note(
                    row_remarks, f"原文数量: {raw_quantity}"
                )
                warnings.append(
                    "数量セルに複数の数値があるため、先頭の数値を数量に使用しました。"
                )
            if item_cell_multiline and not item_cell_split:
                warnings.append(
                    "名称セルが複数行です。複数品目が結合されていないか確認してください。"
                )

            records.append(
                QuantityRecord(
                    source_format="pdf",
                    extractor="pdf_vector_tables",
                    source_sheet=page_title or f"PDF p.{page_number}",
                    source_page=page_number,
                    source_row=row_index + 1,
                    section="本工事費",
                    work_type=current_work_type,
                    category=current_category,
                    item_name=item_name,
                    specification=specification,
                    unit=unit,
                    quantity=quantity,
                    remarks=row_remarks,
                    source_reference=(
                        f"PDF p.{page_number} ({label})"
                        if is_grouped_quantity
                        else f"PDF p.{page_number}"
                    ),
                    extraction_status="PDF_TEXT_REVIEW_REQUIRED",
                    extraction_confidence=0.9 if is_grouped_quantity else 0.97,
                    extraction_warnings=warnings,
                )
            )
    return records


def _extract_pdf_vector_tables(source_path: Path) -> ExtractionResult | None:
    if pdfplumber is None:
        return None

    records: list[QuantityRecord] = []
    selected_pages: list[int] = []
    selected_tables: list[dict[str, Any]] = []
    document_warnings = [
        "このPDFは埋め込み文字を優先して抽出しました。確認用CSVを原本と照合してください。"
    ]
    started = False
    pages_without_table = 0

    try:
        with pdfplumber.open(source_path) as document:
            page_count = len(document.pages)
            for page_number, page in enumerate(document.pages, start=1):
                page_text = page.extract_text() or ""
                if started and _is_pdf_detail_section(page_text):
                    break

                page_tables: list[
                    tuple[list[list[object]], _PdfVectorTableSchema]
                ] = []
                for table in page.extract_tables():
                    schema = _pdf_vector_table_schema(table)
                    if schema is not None:
                        page_tables.append((table, schema))

                if not started:
                    if not page_tables or not _is_pdf_summary_hint(page_text):
                        continue
                    started = True

                if not page_tables:
                    pages_without_table += 1
                    if pages_without_table >= 3:
                        break
                    continue
                pages_without_table = 0

                page_title, scope = _pdf_vector_title_and_scope(page_text)
                for table, schema in page_tables:
                    table_records = _records_from_pdf_vector_table(
                        table,
                        schema,
                        page_number,
                        page_title,
                        scope,
                    )
                    if not table_records:
                        continue
                    records.extend(table_records)
                    if page_number not in selected_pages:
                        selected_pages.append(page_number)
                    selected_tables.append(
                        {
                            "page": page_number,
                            "title": page_title,
                            "layout": schema.layout,
                            "header_row": schema.header_row + 1,
                            "columns": len(table[0]) if table else 0,
                            "quantity_labels": list(schema.quantity_labels),
                            "record_count": len(table_records),
                        }
                    )
    except Exception:
        return None

    if not records:
        return None
    if any(len(table["quantity_labels"]) > 1 for table in selected_tables):
        document_warnings.append(
            "複数の数量列は合計列を優先し、合計が無い表は列別明細として保持しました。"
        )
    return ExtractionResult(
        source_path=str(source_path.resolve()),
        source_format="pdf",
        extractor="pdf_vector_tables",
        project_name=_project_from_filename(source_path),
        records=records,
        document_warnings=document_warnings,
        metadata={
            "page_count": page_count,
            "selected_pages": selected_pages,
            "selected_tables": selected_tables,
        },
    )


def _resolve_pdftoppm() -> Path | None:
    executable = shutil.which("pdftoppm")
    if executable:
        return Path(executable)
    bundled = (
        Path.home()
        / ".cache"
        / "codex-runtimes"
        / "codex-primary-runtime"
        / "dependencies"
        / "native"
        / "poppler"
        / "Library"
        / "bin"
        / "pdftoppm.exe"
    )
    return bundled if bundled.exists() else None


def _render_pdf_pages(
    source_path: Path, output_directory: Path, dpi: int = 200
) -> list[Path]:
    output_directory.mkdir(parents=True, exist_ok=True)
    ascii_source = output_directory / "source.pdf"
    shutil.copyfile(source_path, ascii_source)

    try:
        import pypdfium2 as pdfium  # type: ignore[import-not-found]
    except ImportError:
        pdfium = None

    if pdfium is not None:
        document = pdfium.PdfDocument(str(ascii_source))
        pages: list[Path] = []
        try:
            for page_index in range(len(document)):
                page = document[page_index]
                bitmap = page.render(scale=dpi / 72)
                image = bitmap.to_pil()
                page_path = output_directory / f"page-{page_index + 1:03d}.png"
                image.save(page_path)
                pages.append(page_path)
                image.close()
                bitmap.close()
                page.close()
        finally:
            document.close()
        return pages

    pdftoppm = _resolve_pdftoppm()
    if pdftoppm is None:
        raise QuantityExtractionError(
            "PDF画像化エンジンがありません。setup.ps1を再実行して"
            "pypdfium2をインストールしてください。"
        )
    prefix = output_directory / "page"
    completed = subprocess.run(
        [
            str(pdftoppm),
            "-png",
            "-r",
            str(dpi),
            str(ascii_source),
            str(prefix),
        ],
        capture_output=True,
        timeout=300,
        check=False,
    )
    if completed.returncode != 0:
        message = completed.stderr.decode(errors="replace").strip()
        raise QuantityExtractionError(f"PDFを画像化できませんでした: {message}")
    pages = sorted(output_directory.glob("page-*.png"))
    if not pages:
        raise QuantityExtractionError("PDFを画像化しましたが、ページがありません。")
    return pages


def _powershell_executable() -> str:
    executable = shutil.which("powershell.exe") or shutil.which("powershell")
    if not executable:
        raise QuantityExtractionError(
            "Windows OCRを実行するPowerShellが見つかりません。"
        )
    return executable


def _run_windows_ocr(
    image_directory: Path, language: str = "ja"
) -> dict[str, dict[str, Any]]:
    script = Path(__file__).with_name("windows_ocr.ps1")
    if not script.exists():
        raise QuantityExtractionError(f"OCR補助スクリプトがありません: {script}")
    completed = subprocess.run(
        [
            _powershell_executable(),
            "-NoProfile",
            "-ExecutionPolicy",
            "Bypass",
            "-File",
            str(script),
            "-InputDirectory",
            str(image_directory.resolve()),
            "-Language",
            language,
        ],
        capture_output=True,
        timeout=600,
        check=False,
    )
    stdout = completed.stdout.decode("utf-8-sig", errors="replace").strip()
    stderr = completed.stderr.decode("utf-8", errors="replace").strip()
    if completed.returncode != 0:
        raise QuantityExtractionError(
            "Windows日本語OCRを実行できませんでした: " + (stderr or stdout)
        )
    try:
        payload = json.loads(stdout or "[]")
    except json.JSONDecodeError as error:
        raise QuantityExtractionError(
            f"Windows OCRの結果を解析できませんでした: {error}"
        ) from error
    if isinstance(payload, dict):
        payload = [payload]
    results = {str(entry.get("file", "")): entry for entry in payload}
    errors = [entry["error"] for entry in payload if entry.get("error")]
    if errors and len(errors) == len(payload):
        raise QuantityExtractionError(f"全OCR画像の読取りに失敗しました: {errors[0]}")
    return results


def _make_title_crops(page_paths: list[Path], output_directory: Path) -> None:
    output_directory.mkdir(parents=True, exist_ok=True)
    for page_index, page_path in enumerate(page_paths, start=1):
        with Image.open(page_path) as image:
            crop = image.crop((0, 0, image.width, max(1, int(image.height * 0.28))))
            scale = min(1.5, 2500 / max(crop.width, crop.height))
            if scale > 1:
                crop = crop.resize(
                    (int(crop.width * scale), int(crop.height * scale)),
                    Image.Resampling.LANCZOS,
                )
            canvas = Image.new("RGB", (crop.width + 32, crop.height + 32), "white")
            canvas.paste(crop.convert("RGB"), (16, 16))
            canvas.save(output_directory / f"title-{page_index:03d}.png")
            canvas.close()
            crop.close()


def _normalized_ocr_text(value: object) -> str:
    text = clean_label(value)
    return re.sub(r"\s+", "", text)


def _select_pdf_summary_page(
    source_path: Path,
    page_paths: list[Path],
    workspace: Path,
) -> tuple[int, list[str], dict[str, str]]:
    warnings: list[str] = []
    page_text: dict[str, str] = {}
    try:
        reader = PdfReader(str(source_path))
        for page_index, page in enumerate(reader.pages, start=1):
            page_text[str(page_index)] = page.extract_text() or ""
    except Exception:
        page_text = {}

    for page_index in range(1, len(page_paths) + 1):
        normalized = _normalized_ocr_text(page_text.get(str(page_index), ""))
        if "数量総括表" in normalized:
            return page_index, warnings, page_text

    title_directory = workspace / "title-crops"
    _make_title_crops(page_paths, title_directory)
    title_results = _run_windows_ocr(title_directory)
    for page_index in range(1, len(page_paths) + 1):
        filename = f"title-{page_index:03d}.png"
        text = clean_label(title_results.get(filename, {}).get("text", ""))
        page_text[str(page_index)] = text
        normalized = _normalized_ocr_text(text)
        if "数量総括表" in normalized or (
            "数量" in normalized and "総括" in normalized and "集計" not in normalized
        ):
            return page_index, warnings, page_text

    warnings.append(
        "PDF内で「数量総括表」をOCR確認できなかったため、表罫線が最も多い"
        "ページを候補にしました。ページ選択を確認してください。"
    )
    best_page = 1
    best_score = -1
    for page_index, page_path in enumerate(page_paths, start=1):
        try:
            vertical, horizontal = _detect_table_grid(page_path)
        except QuantityExtractionError:
            continue
        score = len(vertical) * len(horizontal)
        if score > best_score:
            best_page = page_index
            best_score = score
    return best_page, warnings, page_text


def _group_projection_lines(values: list[int], threshold: int) -> list[int]:
    indices = [index for index, value in enumerate(values) if value >= threshold]
    if not indices:
        return []
    groups: list[list[int]] = [[indices[0]]]
    for index in indices[1:]:
        if index == groups[-1][-1] + 1:
            groups[-1].append(index)
        else:
            groups.append([index])
    return [round(sum(group) / len(group)) for group in groups]


def _dense_line_run(lines: list[int], maximum_gap: int) -> list[int]:
    if not lines:
        return []
    runs: list[list[int]] = [[lines[0]]]
    for line in lines[1:]:
        if line - runs[-1][-1] <= maximum_gap:
            runs[-1].append(line)
        else:
            runs.append([line])
    return max(runs, key=len)


def _projection(image: Image.Image, axis: str) -> list[int]:
    if axis == "rows":
        projected = image.resize((1, image.height), Image.Resampling.BOX)
    else:
        projected = image.resize((image.width, 1), Image.Resampling.BOX)
    return list(projected.tobytes())


def _detect_table_grid(page_path: Path) -> tuple[list[int], list[int]]:
    with Image.open(page_path) as source:
        gray = ImageOps.grayscale(source)
        dark = gray.point(lambda pixel: 255 if pixel < 165 else 0)
        page_width, page_height = dark.size

        horizontal: list[int] = []
        for density in (0.55, 0.42, 0.3):
            candidates = _group_projection_lines(
                _projection(dark, "rows"), round(255 * density)
            )
            candidates = _dense_line_run(
                candidates, maximum_gap=max(100, round(page_height * 0.07))
            )
            if len(candidates) >= 3:
                horizontal = candidates
                break
        if len(horizontal) < 3:
            raise QuantityExtractionError("表の横罫線を検出できませんでした。")

        top = max(0, horizontal[0] - 2)
        bottom = min(page_height, horizontal[-1] + 3)
        table_height = dark.crop((0, top, page_width, bottom))
        vertical: list[int] = []
        for density in (0.65, 0.5, 0.35):
            candidates = _group_projection_lines(
                _projection(table_height, "columns"), round(255 * density)
            )
            if 4 <= len(candidates) <= 16:
                vertical = candidates
                break
        table_height.close()
        if len(vertical) < 4:
            raise QuantityExtractionError("表の縦罫線を検出できませんでした。")

        left = max(0, vertical[0] - 2)
        right = min(page_width, vertical[-1] + 3)
        table_width = dark.crop((left, 0, right, page_height))
        for density in (0.65, 0.5, 0.35):
            candidates = _group_projection_lines(
                _projection(table_width, "rows"), round(255 * density)
            )
            candidates = _dense_line_run(
                candidates, maximum_gap=max(100, round(page_height * 0.07))
            )
            if len(candidates) >= 3:
                horizontal = candidates
                break
        table_width.close()

    if len(horizontal) < 3 or len(vertical) < 4:
        raise QuantityExtractionError("数量表の罫線構造を確定できませんでした。")
    return vertical, horizontal


def _cell_has_ink(image: Image.Image) -> bool:
    gray = ImageOps.grayscale(image)
    histogram = gray.histogram()
    dark_pixels = sum(histogram[:200])
    return dark_pixels >= max(6, round(gray.width * gray.height * 0.0008))


def _make_cell_crops(
    page_path: Path,
    vertical: list[int],
    horizontal: list[int],
    output_directory: Path,
) -> None:
    output_directory.mkdir(parents=True, exist_ok=True)
    with Image.open(page_path) as page:
        for row in range(len(horizontal) - 1):
            for column in range(len(vertical) - 1):
                left = vertical[column] + 3
                top = horizontal[row] + 3
                right = vertical[column + 1] - 3
                bottom = horizontal[row + 1] - 3
                if right <= left or bottom <= top:
                    continue
                crop = page.crop((left, top, right, bottom)).convert("RGB")
                if not _cell_has_ink(crop):
                    crop.close()
                    continue
                scale = min(3.0, 2300 / max(crop.width, crop.height))
                if scale > 1:
                    resized = crop.resize(
                        (max(1, round(crop.width * scale)), max(1, round(crop.height * scale))),
                        Image.Resampling.LANCZOS,
                    )
                    crop.close()
                    crop = resized
                canvas = Image.new("RGB", (crop.width + 32, crop.height + 32), "white")
                canvas.paste(crop, (16, 16))
                canvas.save(output_directory / f"r{row:03d}_c{column:03d}.png")
                canvas.close()
                crop.close()


def _make_row_crops(
    page_path: Path,
    vertical: list[int],
    horizontal: list[int],
    output_directory: Path,
) -> dict[str, dict[str, float]]:
    output_directory.mkdir(parents=True, exist_ok=True)
    transforms: dict[str, dict[str, float]] = {}
    with Image.open(page_path) as page:
        for row in range(len(horizontal) - 1):
            left = vertical[0] + 3
            top = horizontal[row] + 3
            right = vertical[-1] - 3
            bottom = horizontal[row + 1] - 3
            if right <= left or bottom <= top:
                continue
            crop = page.crop((left, top, right, bottom)).convert("RGB")
            if not _cell_has_ink(crop):
                crop.close()
                continue
            scale = min(1.5, 2300 / max(crop.width, crop.height))
            if scale > 1:
                resized = crop.resize(
                    (max(1, round(crop.width * scale)), max(1, round(crop.height * scale))),
                    Image.Resampling.LANCZOS,
                )
                crop.close()
                crop = resized
            filename = f"row-{row:03d}.png"
            crop.save(output_directory / filename)
            transforms[filename] = {"left": float(left), "scale": float(scale)}
            crop.close()
    return transforms


def _make_repeated_cell_crops(
    cell_directory: Path,
    initial_results: dict[str, dict[str, Any]],
    output_directory: Path,
) -> None:
    output_directory.mkdir(parents=True, exist_ok=True)
    for source_path in sorted(cell_directory.glob("*.png")):
        if clean_text(initial_results.get(source_path.name, {}).get("text", "")):
            continue
        with Image.open(source_path) as source:
            gray = ImageOps.grayscale(source)
            ink = ImageOps.invert(gray).point(lambda pixel: 255 if pixel > 30 else 0)
            bounding_box = ink.getbbox()
            ink.close()
            if bounding_box is None:
                gray.close()
                continue
            glyph = gray.crop(bounding_box)
            gray.close()
            gap = max(5, round(glyph.width * 0.08))
            repeat_count = 5
            canvas = Image.new(
                "L",
                (
                    glyph.width * repeat_count + gap * (repeat_count + 1),
                    glyph.height + gap * 2,
                ),
                255,
            )
            for index in range(repeat_count):
                canvas.paste(glyph, (gap + index * (glyph.width + gap), gap))
            glyph.close()
            scale = min(2.0, 2300 / max(canvas.width, canvas.height))
            if scale > 1:
                resized = canvas.resize(
                    (round(canvas.width * scale), round(canvas.height * scale)),
                    Image.Resampling.LANCZOS,
                )
                canvas.close()
                canvas = resized
            canvas.save(output_directory / source_path.name)
            canvas.close()


def _decode_repeated_ocr(result: dict[str, Any]) -> str:
    words = [
        clean_text(word.get("text", ""))
        for line in result.get("lines", [])
        for word in line.get("words", [])
        if clean_text(word.get("text", ""))
    ]
    if len(words) >= 3 and len(set(words)) == 1:
        return words[0]
    compact = re.sub(r"\s+", "", clean_text(result.get("text", "")))
    for repeat_count in (5, 4, 3):
        if compact and len(compact) % repeat_count == 0:
            part_length = len(compact) // repeat_count
            parts = [
                compact[index * part_length : (index + 1) * part_length]
                for index in range(repeat_count)
            ]
            if len(set(parts)) == 1:
                return parts[0]
    if len(words) == 1:
        # The retry image contains only copies of one source glyph. Windows OCR
        # sometimes reports one copy even when all five are visible.
        return words[0]
    return ""


def _merge_repeated_ocr(
    initial_results: dict[str, dict[str, Any]],
    repeated_results: dict[str, dict[str, Any]],
) -> None:
    for filename, result in repeated_results.items():
        recovered = _decode_repeated_ocr(result)
        if recovered and not clean_text(initial_results.get(filename, {}).get("text", "")):
            initial_results.setdefault(filename, {})["text"] = recovered


def _ocr_matrix(
    cell_results: dict[str, dict[str, Any]], row_count: int, column_count: int
) -> list[list[str]]:
    rows: list[list[str]] = []
    for row in range(row_count):
        values: list[str] = []
        for column in range(column_count):
            filename = f"r{row:03d}_c{column:03d}.png"
            text = cell_results.get(filename, {}).get("text", "")
            values.append(clean_label(text).replace("\n", " "))
        rows.append(values)
    return rows


def _merge_row_ocr(
    rows: list[list[str]],
    row_results: dict[str, dict[str, Any]],
    transforms: dict[str, dict[str, float]],
    vertical: list[int],
) -> None:
    for filename, transform in transforms.items():
        match = re.fullmatch(r"row-(\d+)\.png", filename)
        if not match:
            continue
        row_index = int(match.group(1))
        if row_index >= len(rows):
            continue
        words_by_column: list[list[tuple[float, float, str]]] = [
            [] for _ in range(len(vertical) - 1)
        ]
        result = row_results.get(filename, {})
        for line_index, line in enumerate(result.get("lines", [])):
            for word in line.get("words", []):
                center = float(word.get("x", 0)) + float(word.get("width", 0)) / 2
                original_x = transform["left"] + center / transform["scale"]
                for column in range(len(vertical) - 1):
                    if vertical[column] <= original_x <= vertical[column + 1]:
                        words_by_column[column].append(
                            (
                                float(line_index),
                                original_x,
                                str(word.get("text", "")),
                            )
                        )
                        break
        for column, words in enumerate(words_by_column):
            if not words:
                continue
            candidate = clean_label(
                " ".join(text for _, _, text in sorted(words))
            )
            existing = clean_label(rows[row_index][column])
            if not existing:
                rows[row_index][column] = candidate
                continue
            existing_compact = re.sub(r"\s+", "", existing)
            candidate_compact = re.sub(r"\s+", "", candidate)
            similarity = SequenceMatcher(
                None, existing_compact, candidate_compact
            ).ratio()
            has_better_measurement = bool(
                re.search(r"(?:L=|\d+(?:\.\d+)?mm)", candidate_compact)
                and not re.search(r"(?:L=|\d+(?:\.\d+)?mm)", existing_compact)
            )
            if has_better_measurement and similarity >= 0.35:
                millimeter = re.search(r"(\d+(?:\.\d+)?)mm", candidate_compact)
                if "L=" in candidate_compact and "L=" not in existing_compact:
                    rows[row_index][column] = candidate
                elif millimeter and re.search(
                    rf"{re.escape(millimeter.group(1))}(?!mm)", existing_compact
                ):
                    rows[row_index][column] = re.sub(
                        rf"{re.escape(millimeter.group(1))}(?!\s*mm)",
                        f"{millimeter.group(1)}mm",
                        existing,
                        count=1,
                    )
                else:
                    rows[row_index][column] = candidate


def _correct_common_ocr_errors(value: str) -> str:
    text = clean_label(value)
    compact = re.sub(r"\s+", "", text)
    if re.fullmatch(r"シー(?:ノ)?レ(?:卞オ|本オ|材)", compact):
        return "シール材"
    replacements = {
        "揚カ": "揚力",
        "べース": "ベース",
        "rn3": "m3",
        "rn2": "m2",
        "r n 3": "m3",
        "r n 2": "m2",
    }
    for old, new in replacements.items():
        text = text.replace(old, new)
    return text


UNIT_GLYPH_CANDIDATES = (
    "m",
    "L",
    "t",
    "本",
    "基",
    "組",
    "式",
    "枚",
    "台",
    "個",
    "人",
    "回",
)


def _normalized_ink(image: Image.Image, size: int = 64) -> tuple[list[int], float]:
    gray = ImageOps.grayscale(image)
    mask = ImageOps.invert(gray).point(lambda pixel: 255 if pixel > 55 else 0)
    bounding_box = mask.getbbox()
    gray.close()
    if bounding_box is None:
        mask.close()
        return [], 0.0
    glyph = mask.crop(bounding_box)
    mask.close()
    aspect = glyph.width / max(1, glyph.height)
    scale = min((size - 8) / glyph.width, (size - 8) / glyph.height)
    resized = glyph.resize(
        (max(1, round(glyph.width * scale)), max(1, round(glyph.height * scale))),
        Image.Resampling.LANCZOS,
    )
    glyph.close()
    canvas = Image.new("L", (size, size), 0)
    canvas.paste(
        resized,
        ((size - resized.width) // 2, (size - resized.height) // 2),
    )
    resized.close()
    vector = [1 if pixel >= 96 else 0 for pixel in canvas.tobytes()]
    canvas.close()
    return vector, aspect


def _unit_font_paths() -> list[Path]:
    font_directory = Path("C:/Windows/Fonts")
    candidates = [
        font_directory / "msmincho.ttc",
        font_directory / "yumin.ttf",
        font_directory / "msgothic.ttc",
        font_directory / "YuGothM.ttc",
    ]
    return [path for path in candidates if path.exists()]


def _unit_template(character: str, font_path: Path) -> tuple[list[int], float]:
    image = Image.new("L", (180, 180), 255)
    font = ImageFont.truetype(str(font_path), 120)
    ImageDraw.Draw(image).text((20, 5), character, font=font, fill=0)
    normalized = _normalized_ink(image)
    image.close()
    return normalized


def _binary_dice(first: list[int], second: list[int]) -> float:
    if not first or not second:
        return 0.0
    first_count = sum(first)
    second_count = sum(second)
    if not first_count or not second_count:
        return 0.0
    intersection = sum(a and b for a, b in zip(first, second))
    return 2 * intersection / (first_count + second_count)


def _classify_unit_glyph(image_path: Path) -> tuple[str, float, float]:
    with Image.open(image_path) as image:
        source_vector, source_aspect = _normalized_ink(image)
    if not source_vector:
        return "", 0.0, 0.0
    scores: dict[str, float] = {}
    for character in UNIT_GLYPH_CANDIDATES:
        best = 0.0
        for font_path in _unit_font_paths():
            template_vector, template_aspect = _unit_template(character, font_path)
            shape_score = _binary_dice(source_vector, template_vector)
            aspect_ratio = min(source_aspect, template_aspect) / max(
                source_aspect, template_aspect
            )
            best = max(best, shape_score * (0.75 + 0.25 * aspect_ratio))
        scores[character] = best
    ranked = sorted(scores.items(), key=lambda item: item[1], reverse=True)
    if not ranked:
        return "", 0.0, 0.0
    best_character, best_score = ranked[0]
    runner_up = ranked[1][1] if len(ranked) > 1 else 0.0
    return best_character, best_score, best_score - runner_up


def _fill_missing_unit_glyphs(
    rows: list[list[str]],
    unit_column: int,
    header_row: int,
    cell_directory: Path,
) -> int:
    recovered = 0
    for row_index in range(header_row + 1, len(rows)):
        if clean_text(rows[row_index][unit_column]):
            continue
        image_path = cell_directory / f"r{row_index:03d}_c{unit_column:03d}.png"
        if not image_path.exists():
            continue
        character, score, margin = _classify_unit_glyph(image_path)
        if score >= 0.62 and margin >= 0.06:
            rows[row_index][unit_column] = character
            recovered += 1
    return recovered


def _pdf_header_columns(
    rows: list[list[str]], column_count: int
) -> tuple[int, dict[str, int], list[str]]:
    warnings: list[str] = []
    best_row = 0
    best_columns: dict[str, int] = {}
    for row_index, values in enumerate(rows[: min(4, len(rows))]):
        columns: dict[str, int] = {}
        for column, value in enumerate(values):
            role = _header_role(value)
            if role and role not in columns:
                columns[role] = column
        if len(columns) > len(best_columns):
            best_row = row_index
            best_columns = columns

    if {"item_name", "unit", "quantity"}.issubset(best_columns):
        return best_row, best_columns, warnings
    if column_count == 7:
        warnings.append(
            "PDFの見出しOCRが不完全なため、珠洲市数量総括表の7列順を使用しました。"
        )
        return 0, {
            "work_type": 0,
            "category": 1,
            "item_name": 2,
            "specification": 3,
            "unit": 4,
            "quantity": 5,
            "remarks": 6,
        }, warnings
    if column_count == 8:
        warnings.append(
            "PDFの見出しOCRが不完全なため、費目を含む8列順を使用しました。"
        )
        return 0, {
            "section": 0,
            "work_type": 1,
            "category": 2,
            "item_name": 3,
            "specification": 4,
            "unit": 5,
            "quantity": 6,
            "remarks": 7,
        }, warnings
    raise QuantityExtractionError(
        f"PDF表の列見出しを確定できませんでした ({column_count}列)。"
    )


def _parse_ocr_quantity(value: str) -> float | int | None:
    direct = parse_quantity(value)
    if direct is not None:
        return direct
    text = clean_text(value).replace(" ", "").replace(",", "")
    text = text.replace("O", "0").replace("o", "0").replace("．", ".")
    match = re.fullmatch(r"[-+]?\d+(?:\.\d+)?", text)
    return parse_quantity(match.group(0)) if match else None


def _pdf_value(row: list[str], columns: dict[str, int], role: str) -> str:
    column = columns.get(role)
    return (
        _correct_common_ocr_errors(row[column])
        if column is not None and column < len(row)
        else ""
    )


def _records_from_pdf_rows(
    rows: list[list[str]],
    header_row: int,
    columns: dict[str, int],
    page_number: int,
) -> list[QuantityRecord]:
    current = {
        "section": "本工事費",
        "work_type": "",
        "category": "",
        "item_name": "",
        "unit": "",
    }
    records: list[QuantityRecord] = []
    for row_index in range(header_row + 1, len(rows)):
        row = rows[row_index]
        explicit = {
            role: _pdf_value(row, columns, role)
            for role in (
                "section",
                "work_type",
                "category",
                "item_name",
                "specification",
                "unit",
                "remarks",
            )
        }
        if not any(explicit.values()):
            continue
        if _is_total_row(*explicit.values()):
            continue

        if explicit["section"]:
            current["section"] = explicit["section"]
        if explicit["work_type"]:
            current["work_type"] = explicit["work_type"]
            current["category"] = ""
            current["item_name"] = ""
        if explicit["category"]:
            current["category"] = explicit["category"]
            current["item_name"] = ""
        if explicit["item_name"]:
            current["item_name"] = explicit["item_name"]
        if explicit["unit"]:
            current["unit"] = explicit["unit"]

        quantity = _parse_ocr_quantity(_pdf_value(row, columns, "quantity"))
        item_name = explicit["item_name"] or current["item_name"]
        unit = explicit["unit"]
        if not unit and not explicit["item_name"] and (
            explicit["specification"] or quantity is not None
        ):
            unit = current["unit"]
        if not item_name:
            continue
        if quantity is None and not unit:
            continue

        warnings = [
            "PDF OCR行です。原本の名称・規格・単位・数量を確認してください。"
        ]
        if quantity is None:
            warnings.append("数量OCRを確定できませんでした。")
        if not unit:
            warnings.append("単位OCRを確定できませんでした。")
        records.append(
            QuantityRecord(
                source_format="pdf",
                extractor="pdf_grid_windows_ocr",
                source_sheet=f"PDF p.{page_number}",
                source_page=page_number,
                source_row=row_index + 1,
                section=current["section"],
                work_type=current["work_type"],
                category=current["category"],
                item_name=item_name,
                specification=explicit["specification"],
                unit=unit,
                quantity=quantity,
                remarks=explicit["remarks"],
                source_reference=f"PDF p.{page_number}",
                extraction_status="OCR_REVIEW_REQUIRED",
                extraction_confidence=0.7 if quantity is not None and unit else 0.45,
                extraction_warnings=warnings,
            )
        )
    return records


def _extract_pdf_in_workspace(
    source_path: Path, workspace: Path, preserve_workspace: bool
) -> ExtractionResult:
    page_directory = workspace / "pages"
    page_paths = _render_pdf_pages(source_path, page_directory)
    page_number, warnings, page_text = _select_pdf_summary_page(
        source_path, page_paths, workspace
    )
    selected_page = page_paths[page_number - 1]
    vertical, horizontal = _detect_table_grid(selected_page)
    cell_directory = workspace / "cells"
    _make_cell_crops(selected_page, vertical, horizontal, cell_directory)
    cell_results = _run_windows_ocr(cell_directory)
    repeated_directory = workspace / "repeated-cells"
    _make_repeated_cell_crops(cell_directory, cell_results, repeated_directory)
    if any(repeated_directory.glob("*.png")):
        repeated_results = _run_windows_ocr(repeated_directory)
        _merge_repeated_ocr(cell_results, repeated_results)
        repeated_english_results = _run_windows_ocr(
            repeated_directory, language="en-US"
        )
        _merge_repeated_ocr(cell_results, repeated_english_results)
    rows = _ocr_matrix(
        cell_results,
        row_count=len(horizontal) - 1,
        column_count=len(vertical) - 1,
    )
    row_directory = workspace / "rows"
    row_transforms = _make_row_crops(
        selected_page, vertical, horizontal, row_directory
    )
    row_results = _run_windows_ocr(row_directory)
    _merge_row_ocr(rows, row_results, row_transforms, vertical)
    header_row, columns, header_warnings = _pdf_header_columns(
        rows, len(vertical) - 1
    )
    warnings.extend(header_warnings)
    recovered_units = 0
    if "unit" in columns:
        recovered_units = _fill_missing_unit_glyphs(
            rows, columns["unit"], header_row, cell_directory
        )
    if recovered_units:
        warnings.append(
            f"単文字単位を画像字形で{recovered_units}件補完しました。該当単位を確認してください。"
        )
    warnings.append(
        "PDFは画像OCRで抽出しました。GAIA出力前にnormalized.csvを原本と照合してください。"
    )
    records = _records_from_pdf_rows(rows, header_row, columns, page_number)
    if not records:
        raise QuantityExtractionError(
            f"PDF p.{page_number}の表から数量明細を抽出できませんでした。"
        )

    metadata: dict[str, Any] = {
        "selected_page": page_number,
        "page_count": len(page_paths),
        "grid": {
            "columns": len(vertical) - 1,
            "rows": len(horizontal) - 1,
            "vertical_lines": vertical,
            "horizontal_lines": horizontal,
        },
        "page_title_ocr": page_text,
        "header_columns": columns,
    }
    if preserve_workspace:
        metadata["ocr_workspace"] = str(workspace.resolve())
    return ExtractionResult(
        source_path=str(source_path.resolve()),
        source_format="pdf",
        extractor="pdf_grid_windows_ocr",
        project_name=_project_from_filename(source_path),
        records=records,
        document_warnings=warnings,
        metadata=metadata,
    )


def extract_pdf_quantity_source(
    source_path: Path, work_dir: Path | None = None
) -> ExtractionResult:
    vector_result = _extract_pdf_vector_tables(source_path)
    if vector_result is not None:
        return vector_result
    if work_dir is not None:
        digest = hashlib.sha1(str(source_path.resolve()).encode("utf-8")).hexdigest()[:8]
        workspace = work_dir / f"{source_path.stem[:30]}_{digest}"
        workspace.mkdir(parents=True, exist_ok=True)
        return _extract_pdf_in_workspace(source_path, workspace, True)
    with tempfile.TemporaryDirectory(prefix="gaia-quantity-pdf-") as directory:
        return _extract_pdf_in_workspace(source_path, Path(directory), False)


def extract_quantity_source(
    source_path: Path, work_dir: Path | None = None
) -> ExtractionResult:
    source_path = Path(source_path)
    if not source_path.exists():
        raise FileNotFoundError(f"数量計算書がありません: {source_path}")
    suffix = source_path.suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        return extract_excel_quantity_source(source_path)
    if suffix == ".pdf":
        return extract_pdf_quantity_source(source_path, work_dir=work_dir)
    raise QuantityExtractionError(
        f"未対応のファイル形式です: {suffix or '(拡張子なし)'}。"
        "現在は .xlsx, .xlsm, .pdf に対応しています。"
    )
