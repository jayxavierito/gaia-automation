import unittest
from pathlib import Path

from openpyxl import Workbook

from gaia_catalog import (
    GaiaCatalogEntry,
    extract_gaia_catalog,
    match_gaia_catalog,
)
from gaia_converter import (
    BoqItem,
    apply_gaia_catalog,
    build_output_blocks,
    classify_schema_item,
    write_block,
)


class GaiaCatalogTests(unittest.TestCase):
    def make_item(self, specification="再生密粒度アスコン t=5cm"):
        return BoqItem(
            source_sheet="数量総括",
            source_row=4,
            section="本工事費",
            work_type="1. 舗装工",
            category="舗装打換え工",
            item_name="表層",
            specification=specification,
            unit="m2",
            quantity=100,
            remarks="原本の不要な備考",
            standard="",
            daily_output="",
            setting="",
            notes="",
            source_reference="",
        )

    def make_entry(
        self,
        *,
        specification="再生密粒度アスコン t=5cm",
        code="CB410260",
        source_row=16,
    ):
        return GaiaCatalogEntry(
            level1="道路災害3911号",
            level2="舗装工",
            level3="舗装打換え工",
            item_name="表層",
            specification=specification,
            unit="m2",
            code=code,
            source_file="accepted.xlsx",
            source_row=source_row,
        )

    def test_extracts_gaia_hierarchy_and_item_code_from_accepted_layout(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "本工事費内訳表"
        sheet["A4"] = "道路災害3911号"
        sheet["M5"] = "費目行"
        sheet["B8"] = "舗装工"
        sheet["M9"] = "工種行"
        sheet["C12"] = "舗装打換え工"
        sheet["M13"] = "種別行"
        sheet["D16"] = "表層"
        sheet["H16"] = 100
        sheet["J16"] = "m2"
        sheet["M16"] = "[CB410260]"
        sheet["D17"] = "再生密粒度アスコン t=5cm"

        path = Path.cwd() / "test_gaia_accepted.xlsx"
        try:
            workbook.save(path)
            entries = extract_gaia_catalog(path)
        finally:
            path.unlink(missing_ok=True)

        self.assertEqual(len(entries), 1)
        self.assertEqual(entries[0].level2, "舗装工")
        self.assertEqual(entries[0].level3, "舗装打換え工")
        self.assertEqual(entries[0].code, "CB410260")

    def test_extracts_deep_hierarchy_without_cataloging_group_rows(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "本工事費内訳表"
        sheet["A4"] = "道路災害3911号"
        sheet["M5"] = "費目行"
        sheet["B8"] = "舗装工"
        sheet["M9"] = "工種行"
        sheet["C12"] = "舗装打換え工"
        sheet["M13"] = "種別行"
        sheet["D16"] = "橋面防水工"
        sheet["H16"] = 1
        sheet["J16"] = "式"
        sheet["M17"] = "細別行"
        sheet["E20"] = "橋面防水"
        sheet["H20"] = 1
        sheet["J20"] = "式"
        sheet["M21"] = "規格行"
        sheet["F24"] = "シート系防水"
        sheet["H24"] = 463
        sheet["J24"] = "m2"
        sheet["M24"] = "[CB410260]"
        sheet["D25"] = "新設・全面接着"
        sheet["M25"] = "施工 第 0-189号表"

        path = Path.cwd() / "test_gaia_deep_hierarchy.xlsx"
        try:
            workbook.save(path)
            entries = extract_gaia_catalog(path)
        finally:
            path.unlink(missing_ok=True)

        self.assertEqual(len(entries), 1)
        self.assertEqual(entries[0].level4, "橋面防水工")
        self.assertEqual(entries[0].level5, "橋面防水")
        self.assertEqual(entries[0].item_name, "シート系防水")
        self.assertEqual(entries[0].item_column, 6)
        self.assertEqual(entries[0].reference_type, "施工")
        self.assertEqual(entries[0].reference_text, "施工 第 0-189号表")

    def test_exact_name_unit_and_condition_can_supply_safe_code(self):
        match = match_gaia_catalog(
            [self.make_entry()],
            item_name="表層",
            specification="再生密粒度アスコン t=5cm",
            unit="m2",
            work_type="舗装工",
            category="舗装打換え工",
        )

        self.assertEqual(match.status, "EXACT_CODE")
        self.assertTrue(match.path_safe)
        self.assertTrue(match.code_safe)

    def test_ambiguous_codes_are_not_written_without_condition_evidence(self):
        entries = [
            self.make_entry(
                specification="密粒度アスコン t=5cm",
                code="CB410260",
                source_row=16,
            ),
            self.make_entry(
                specification="密粒度アスコン t=7cm",
                code="CB410270",
                source_row=20,
            ),
        ]
        match = match_gaia_catalog(
            entries,
            item_name="表層",
            specification="",
            unit="m2",
            work_type="舗装工",
            category="舗装打換え工",
        )

        self.assertTrue(match.path_safe)
        self.assertFalse(match.code_safe)

    def test_explicit_source_hierarchy_is_not_replaced_by_catalog_branch(self):
        item = self.make_item()
        entry = self.make_entry()
        entry = GaiaCatalogEntry(
            **{
                **entry.__dict__,
                "level2": "別の工種",
                "level3": "別の種別",
            }
        )

        apply_gaia_catalog([item], [entry])

        self.assertEqual(item.output_level2, "舗装工")
        self.assertEqual(item.output_level3, "舗装打換え工")

    def test_generic_code_requires_specification_or_context_support(self):
        entry = GaiaCatalogEntry(
            level1="道路災害3911号",
            level2="下部工",
            level3="橋台躯体工",
            item_name="アンカー工",
            specification="",
            unit="本",
            code="CB435940",
            source_file="accepted.xlsx",
            source_row=100,
        )
        match = match_gaia_catalog(
            [entry],
            item_name="アンカー工",
            specification="",
            unit="本",
            work_type="胸壁再構築工",
            category="胸壁打設工",
        )

        self.assertFalse(match.code_safe)

    def test_blank_hierarchy_rows_are_not_treated_as_siblings(self):
        matched = self.make_item()
        matched.work_type = ""
        matched.category = ""
        unmatched = self.make_item()
        unmatched.source_row = 8
        unmatched.work_type = ""
        unmatched.category = ""
        unmatched.item_name = "カタログにない別の作業"
        unmatched.unit = "箇所"

        apply_gaia_catalog([matched, unmatched], [self.make_entry()])

        self.assertEqual(matched.output_level2, "舗装工")
        self.assertEqual(matched.output_level3, "舗装打換え工")
        self.assertEqual(unmatched.output_level2, "工種未確認")
        self.assertEqual(unmatched.output_level3, "種別未確認")
        self.assertEqual(unmatched.catalog_status, "NOT_FOUND")

    def test_output_uses_exact_gaia_markers_and_only_required_item_data(self):
        item = self.make_item()
        apply_gaia_catalog([item], [self.make_entry()])
        classify_schema_item(item)
        blocks = build_output_blocks([item])

        workbook = Workbook()
        sheet = workbook.active
        for index, block in enumerate(blocks):
            write_block(sheet, index, block)

        self.assertEqual(sheet["M5"].value, "費目行")
        self.assertEqual(sheet["M9"].value, "工種行")
        self.assertEqual(sheet["M13"].value, "種別行")
        self.assertEqual(sheet["D16"].value, "表層")
        self.assertEqual(sheet["H16"].value, 100)
        self.assertEqual(sheet["J16"].value, "m2")
        self.assertEqual(sheet["M16"].value, "[CB410260]")
        self.assertEqual(sheet["D17"].value, "再生密粒度アスコン t=5cm")
        self.assertIsNone(sheet["M17"].value)
        self.assertIsNone(sheet["M18"].value)
        self.assertIsNone(sheet["M19"].value)

    def test_output_preserves_deep_markers_but_omits_stale_prices_and_references(self):
        item = self.make_item()
        entry = GaiaCatalogEntry(
            level1="道路災害3911号",
            level2="舗装工",
            level3="舗装打換え工",
            item_name="シート系防水",
            specification="新設・全面接着",
            unit="m2",
            code="CB410260",
            source_file="accepted.xlsx",
            source_row=1768,
            level4="橋面防水工",
            level5="橋面防水",
            item_column=6,
            reference_type="施工",
            reference_text="施工 第 0-189号表",
        )
        item.item_name = "シート系防水"
        item.specification = "新設・全面接着"
        item.gaia_condition = item.specification
        apply_gaia_catalog([item], [entry])
        classify_schema_item(item)

        workbook = Workbook()
        sheet = workbook.active
        for index, block in enumerate(build_output_blocks([item])):
            write_block(sheet, index, block)

        self.assertEqual(sheet["D16"].value, "橋面防水工")
        self.assertEqual(sheet["M17"].value, "細別行")
        self.assertEqual(sheet["E20"].value, "橋面防水")
        self.assertEqual(sheet["M21"].value, "規格行")
        self.assertEqual(sheet["F24"].value, "シート系防水")
        self.assertEqual(sheet["D25"].value, "新設・全面接着")
        self.assertEqual(sheet["M24"].value, "[CB410260]")
        self.assertIsNone(sheet["M25"].value)
        for row in (4, 8, 12, 16, 20, 24):
            self.assertIsNone(sheet.cell(row, 11).value)
            self.assertIsNone(sheet.cell(row, 12).value)


if __name__ == "__main__":
    unittest.main()
