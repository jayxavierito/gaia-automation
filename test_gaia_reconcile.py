import csv
import unittest
from pathlib import Path

from openpyxl import Workbook

from gaia_reconcile import (
    load_source_review,
    parse_gaia_review_workbook,
    reconcile_rows,
)


class ReconcileTests(unittest.TestCase):
    def setUp(self):
        self.review_path = Path.cwd() / "test_gaia_review.xlsx"
        self.source_path = Path.cwd() / "test_source_review.csv"

    def tearDown(self):
        self.review_path.unlink(missing_ok=True)
        self.source_path.unlink(missing_ok=True)

    def test_parses_gaia_columns_and_matches_source_row(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "確認単価"
        sheet.cell(7, 2).value = "床掘り"
        sheet.cell(7, 17).value = "土砂 小規模"
        sheet.cell(7, 32).value = "m3"
        sheet.cell(7, 50).value = "単価金額が 0円または 1円です。"
        sheet.cell(7, 68).value = 0
        workbook.create_sheet("設定")
        workbook.save(self.review_path)
        workbook.close()

        fieldnames = [
            "item_name",
            "gaia_item_name",
            "specification",
            "gaia_condition",
            "unit",
            "package_unit",
            "package_unit_status",
            "match_status",
        ]
        with self.source_path.open(
            "w", encoding="utf-8-sig", newline=""
        ) as handle:
            writer = csv.DictWriter(handle, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerow(
                {
                    "item_name": "床掘り",
                    "gaia_item_name": "床掘り",
                    "specification": "土砂 小規模",
                    "gaia_condition": "土砂 小規模",
                    "unit": "m3",
                    "package_unit": "m3",
                    "package_unit_status": "MATCH",
                    "match_status": "PACKAGE_DATE_REVIEW",
                }
            )

        gaia_rows = parse_gaia_review_workbook(self.review_path, "unit")
        source_rows = load_source_review(self.source_path)
        result = reconcile_rows(source_rows, gaia_rows)

        self.assertEqual(len(gaia_rows), 1)
        self.assertEqual(gaia_rows[0].price, "0")
        self.assertEqual(result[0]["gaia_review_status"], "GAIA_UNIT_REVIEW")
        self.assertIn("0円", result[0]["gaia_review_message"])

    def test_keeps_unmatched_gaia_row_for_manual_review(self):
        from gaia_reconcile import GaiaReviewRow

        review = GaiaReviewRow(
            queue="work",
            workbook="review.xlsx",
            gaia_row=7,
            name="表層",
            specification="",
            unit="m2",
            price="0",
            message="再実行してください",
        )
        result = reconcile_rows([], [review])

        self.assertEqual(len(result), 1)
        self.assertEqual(
            result[0]["gaia_review_status"], "UNMATCHED_GAIA_REVIEW_ROW"
        )

    def test_aggregates_multiple_gaia_issues_for_one_source_line(self):
        from gaia_reconcile import GaiaReviewRow

        source = {
            "item_name": "上層路盤(車道・路肩部)",
            "gaia_item_name": "上層路盤(車道・路肩部)",
            "specification": "M-40 t=15cm",
            "gaia_condition": "M-40 t=15cm",
            "unit": "m2",
            "package_unit": "m2",
            "package_unit_status": "MATCH",
        }
        reviews = [
            GaiaReviewRow(
                queue="work",
                workbook="review.xlsx",
                gaia_row=row,
                name="上層路盤(車道・路肩部)",
                specification="M-40 t=15cm",
                unit="m2",
                price="0",
                message=message,
            )
            for row, message in ((7, "表の単価が0円"), (8, "条件選択が一時中止"))
        ]

        result = reconcile_rows([source], reviews)

        self.assertEqual(len(result), 1)
        self.assertEqual(result[0]["gaia_review_status"], "GAIA_WORK_REVIEW")
        self.assertIn("表の単価", result[0]["gaia_review_message"])
        self.assertIn("一時中止", result[0]["gaia_review_message"])


if __name__ == "__main__":
    unittest.main()
