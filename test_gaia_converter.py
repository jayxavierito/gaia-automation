import unittest
from datetime import date
from pathlib import Path

from openpyxl import Workbook

from gaia_catalog import GaiaCatalogEntry
from gaia_converter import (
    BoqItem,
    apply_gaia_catalog,
    build_gaia_candidate,
    build_output_blocks,
    classify_schema_item,
    detect_estimate_date,
    extract_boq_items,
    infer_work_category,
    normalize_hierarchy_label,
    normalize_match_text,
)

ASSETS_DIR = Path(__file__).resolve().parent / "assets"


class ConverterTests(unittest.TestCase):
    def test_extracts_two_row_summary_item(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "設計数量総括表(上部工)"
        sheet["A2"] = "設計書名：テスト橋"
        sheet.append(["工種", "種別", "細別", "規格", "単位", "数量"])
        sheet.append(["舗装工", None, None, None, None, None])
        sheet.append([None, "舗装工", None, None, None, None])
        sheet.append(
            [
                None,
                None,
                "表層(車道・路肩部)",
                "再生密粒度アスコン t=5cm",
                "m2",
                None,
                None,
                "R7_国土交通省土木工事積算基準",
                "2300m2",
                "3.0m超",
                None,
            ]
        )
        sheet.append([None, None, None, None, None, 12.5, None, "P.1442"])

        path = Path.cwd() / "test_source.xlsx"
        try:
            workbook.save(path)
            project_name, items = extract_boq_items(path)
        finally:
            path.unlink(missing_ok=True)

        self.assertEqual(project_name, "テスト橋")
        self.assertEqual(len(items), 1)
        self.assertEqual(items[0].quantity, 12.5)
        self.assertEqual(items[0].work_type, "舗装工")
        self.assertEqual(items[0].category, "舗装工")
        self.assertEqual(items[0].source_reference, "P.1442")

    def test_hierarchy_sequence_prefix_is_removed_without_touching_names(self):
        self.assertEqual(
            normalize_hierarchy_label("1. 下部工補修工"), "下部工補修工"
        )
        self.assertEqual(normalize_hierarchy_label("1号橋補修工"), "1号橋補修工")

    def test_trailing_work_suffix_and_tsumikomi_are_normalized(self):
        self.assertEqual(
            normalize_match_text("排水管設置工"),
            normalize_match_text("排水管設置"),
        )
        self.assertEqual(
            normalize_match_text("積込み(コンクリート殻)"),
            normalize_match_text("積込(コンクリート殻)"),
        )

    def test_detects_estimate_month_next_to_excel_label(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet["A1"] = "積算年月"
        sheet["B1"] = "令和8年6月"
        path = Path.cwd() / "test_estimate_date.xlsx"
        try:
            workbook.save(path)
            detected, source = detect_estimate_date(path)
        finally:
            path.unlink(missing_ok=True)

        self.assertEqual(detected, date(2026, 6, 1))
        self.assertEqual(source, "SOURCE")

    def test_missing_estimate_month_uses_pc_date(self):
        workbook = Workbook()
        workbook.active["A1"] = "数量計算書"
        path = Path.cwd() / "test_no_estimate_date.xlsx"
        try:
            workbook.save(path)
            detected, source = detect_estimate_date(path)
        finally:
            path.unlink(missing_ok=True)

        self.assertEqual(detected, date.today())
        self.assertEqual(source, "PC_DATE")

    def _make_item(self, **overrides) -> BoqItem:
        fields = dict(
            source_sheet="数量総括",
            source_row=4,
            section="本工事費",
            work_type="舗装工",
            category="舗装打換え工",
            item_name="表層",
            specification="再生密粒度アスコン t=5cm",
            unit="m2",
            quantity=100,
            remarks="",
            standard="",
            daily_output="",
            setting="",
            notes="",
            source_reference="",
        )
        fields.update(overrides)
        return BoqItem(**fields)

    def test_invalid_quantity_item_is_excluded_from_output_blocks(self):
        zero_quantity_item = self._make_item(quantity=0)
        apply_gaia_catalog([zero_quantity_item], [])
        classify_schema_item(zero_quantity_item)

        self.assertEqual(zero_quantity_item.match_status, "INVALID_QUANTITY")
        self.assertEqual(build_output_blocks([zero_quantity_item]), [])

    def test_source_review_required_item_never_receives_a_code_even_with_exact_catalog_match(
        self,
    ):
        entry = GaiaCatalogEntry(
            level1="本工事費",
            level2="舗装工",
            level3="舗装打換え工",
            item_name="表層",
            specification="再生密粒度アスコン t=5cm",
            unit="m2",
            code="CB410260",
            source_file="accepted.xlsx",
            source_row=16,
        )
        item = self._make_item(extraction_status="SOURCE_REVIEW_REQUIRED")
        apply_gaia_catalog([item], [entry])
        classify_schema_item(item)

        self.assertEqual(item.match_status, "SOURCE_REVIEW_REQUIRED")
        self.assertEqual(item.gaia_code, "")

    def test_infer_work_category_uses_project_name_and_item_keywords(self):
        bridge_item = self._make_item(
            work_type="橋台工", category="躯体工", item_name="床版工"
        )
        self.assertEqual(
            infer_work_category("〇〇橋補修工事", [bridge_item]), "橋梁工事"
        )

    def test_build_gaia_candidate_writes_hierarchy_into_shipped_template(self):
        template_path = ASSETS_DIR / "gaia_suzu_7sheet_template.xlsx"
        if not template_path.exists():
            self.skipTest("shipped GAIA template is not present in this checkout")

        item = self._make_item()
        apply_gaia_catalog([item], [])
        classify_schema_item(item)

        output_path = Path.cwd() / "test_gaia_candidate.xlsx"
        try:
            block_count = build_gaia_candidate(
                template_path=template_path,
                output_path=output_path,
                project_name="テスト工事",
                items=[item],
                location="テスト地内",
                work_category="道路工事",
                district="珠洲",
                price_date=date(2026, 4, 1),
            )
            self.assertEqual(block_count, 4)  # fee + work + category + item rows

            from openpyxl import load_workbook

            workbook = load_workbook(output_path)
            try:
                main_sheet = workbook["本工事費内訳表"]
                self.assertEqual(main_sheet["D16"].value, "表層")
                self.assertEqual(main_sheet["H16"].value, 100)
                self.assertEqual(main_sheet["J16"].value, "m2")
                self.assertEqual(workbook["鏡"]["D6"].value, "テスト工事")
            finally:
                workbook.close()
        finally:
            output_path.unlink(missing_ok=True)


if __name__ == "__main__":
    unittest.main()
